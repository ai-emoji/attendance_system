"""export.export_details

Xuất "chi tiết chấm công" từ QTableWidget (Shift Attendance - MainContent2) ra Excel.

Cấu trúc header giữ giống Xuất lưới (và dùng chung settings từ export_grid_list_dialog):
- Dòng 1: Tên công ty
- Dòng 2: Địa chỉ
- Dòng 3: Số điện thoại
- Dòng 4: Chi tiết chấm công
- Dòng 5: Từ ngày ... đến ngày ...

Phần bảng:
- Header: tất cả cột (trừ cột checkbox ✅/❌), không phụ thuộc ẩn/hiện.
- Dữ liệu: text đang hiển thị trong QTableWidget.

Cuối file:
- Người tạo danh sách
- Ghi chú
"""

from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
import unicodedata

from export.export_grid_list import CompanyInfo


def export_shift_attendance_details_xlsx(
    *,
    file_path: str,
    company: CompanyInfo,
    from_date_text: str,
    to_date_text: str,
    table,
    row_indexes: list[int] | None = None,
    force_exclude_headers: set[str] | None = None,
    in_out_mode_by_employee_code: dict[str, str | None] | None = None,
    department_text: str | None = None,
    title_text: str | None = None,
    company_name_style: dict | None = None,
    company_address_style: dict | None = None,
    company_phone_style: dict | None = None,
    creator: str | None = None,
    creator_style: dict | None = None,
    note_text: str | None = None,
    note_style: dict | None = None,
) -> tuple[bool, str]:
    path = Path(str(file_path or "").strip())
    if not str(path):
        return False, "Vui lòng chọn đường dẫn file xuất."
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return (
            False,
            "Thiếu thư viện xuất Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
        )

    # Detail export: include all columns except checkbox column 0
    try:
        col_count = int(table.columnCount())
    except Exception:
        return False, "Không thể đọc cấu trúc cột của bảng để xuất."

    excluded_headers = {str(x or "").strip() for x in (force_exclude_headers or set())}

    cols: list[int] = []
    for c in range(col_count):
        if int(c) == 0:
            continue

        # User requirement: export should match what is shown on the table.
        # If a column is hidden in the UI, do not export it.
        try:
            if bool(table.isColumnHidden(int(c))):
                continue
        except Exception:
            pass

        try:
            hi = table.horizontalHeaderItem(int(c))
            ht = "" if hi is None else str(hi.text() or "").strip()
        except Exception:
            ht = ""
        if not ht:
            continue
        if excluded_headers and ht in excluded_headers:
            continue
        cols.append(int(c))

    if not cols:
        return False, "Không có cột để xuất."

    wb = Workbook()
    ws = wb.active
    ws.title = "XuatChiTiet"

    ncols = len(cols)
    grid_ncols = ncols

    def _merge_full_row(row: int) -> None:
        if ncols <= 1:
            return
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

    title_font = Font(bold=True)
    header_font = Font(bold=True)
    grid_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _parse_date_any(s: str) -> date | None:
        t = str(s or "").strip()
        if not t:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(t, fmt).date()
            except Exception:
                continue
        return None

    def _vn_weekday(d: date) -> str:
        # Monday=0 .. Sunday=6
        wd = int(d.weekday())
        if wd == 6:
            return "Chủ nhật"
        return f"Thứ {wd + 2}"

    def _pick_date_cell_format(table_date_samples: list[str]) -> str:
        for s in table_date_samples:
            t = str(s or "").strip()
            if not t:
                continue
            if _parse_date_any(t) is None:
                continue
            if len(t) == 10 and t[4] == "-":
                return "%Y-%m-%d"
            if len(t) == 10 and t[2] == "/":
                return "%d/%m/%Y"
        # default: match DB/work_date format
        return "%Y-%m-%d"

    def _norm_style(
        st: dict | None, *, default_align: str = "center"
    ) -> tuple[Font, Alignment]:
        d = st or {}
        try:
            size = int(d.get("font_size", 13))
        except Exception:
            size = 13
        size = max(8, min(40, size))
        bold = bool(d.get("bold", False))
        italic = bool(d.get("italic", False))
        underline = bool(d.get("underline", False))
        align = str(d.get("align", default_align) or default_align).strip().lower()
        if align not in {"left", "center", "right"}:
            align = default_align
        f = Font(
            size=size,
            bold=bold,
            italic=italic,
            underline=("single" if underline else None),
        )
        a = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return f, a

    # Row 1..5 (merged)
    _merge_full_row(1)
    c1 = ws.cell(
        row=1,
        column=1,
        value=f"Tên công ty : {str(company.name or '').strip()}",
    )
    if company_name_style is not None:
        f, a = _norm_style(company_name_style, default_align="center")
        c1.font = f
        c1.alignment = a
    else:
        c1.font = title_font
        c1.alignment = center

    _merge_full_row(2)
    c2 = ws.cell(
        row=2,
        column=1,
        value=f"Địa chỉ : {str(company.address or '').strip()}",
    )
    if company_address_style is not None:
        f, a = _norm_style(company_address_style, default_align="center")
        c2.font = f
        c2.alignment = a
    else:
        c2.alignment = center

    _merge_full_row(3)
    c3 = ws.cell(
        row=3,
        column=1,
        value=f"Số điện thoại : {str(company.phone or '').strip()}",
    )
    if company_phone_style is not None:
        f, a = _norm_style(company_phone_style, default_align="center")
        c3.font = f
        c3.alignment = a
    else:
        c3.alignment = center

    _merge_full_row(4)
    c4 = ws.cell(row=4, column=1, value="Chi tiết chấm công")
    c4.font = title_font
    c4.alignment = center

    _merge_full_row(5)
    c5 = ws.cell(
        row=5,
        column=1,
        value=f"Từ ngày: {str(from_date_text or '').strip()}    Đến ngày: {str(to_date_text or '').strip()}",
    )
    c5.font = title_font
    c5.alignment = center

    # Parse date range
    from_d = _parse_date_any(from_date_text)
    to_d = _parse_date_any(to_date_text)

    def _norm_symbol_text(v: object | None) -> str:
        s = "" if v is None else str(v)
        s = unicodedata.normalize("NFC", s)
        return s.strip()

    def _is_merge_status_symbol(v: object | None) -> bool:
        # Requirement: when exporting detail, merge V/OFF/Lễ (and commonly 'Le')
        # from IN cell into the corresponding IN/OUT pair.
        t = _norm_symbol_text(v).lower()
        if not t:
            return False
        return t in {"v", "off", "le", "lễ"}

    def _strip_export_symbols(txt: object | None, header: str = "") -> str:
        """Remove UI symbols from exported text (e.g. '4.5 +' -> '4.5', '10 Tr' -> '10').

        We intentionally do NOT sanitize identity columns like employee name/code.
        """
        s = _norm_symbol_text(txt)
        if not s:
            return ""

        # KV/KR are meaningful attendance symbols (missing IN/OUT) and must be exported.
        # They are not a decorative UI suffix like '+', 'Tr', 'Sm'.
        if s.strip().lower() in {"kv", "kr"}:
            return s

        h = str(header or "").strip().lower()
        if h in {
            "stt",
            "mã nv",
            "mã nhân viên",
            "tên nhân viên",
            "ngày",
            "thứ",
            "lịch",
            "ca",
            "schedule",
        }:
            return s

        # Common attendance UI symbols (suffixes) that should be removed from numeric exports.
        # Note: KV/KR are handled above and intentionally not stripped.
        symbol_tokens = {"+", "tr", "sm", "x", "v", "off", "le", "lễ", "đ"}

        parts = s.split()
        if len(parts) >= 2 and parts[-1].strip().lower() in symbol_tokens:
            s = " ".join(parts[:-1]).strip()
            return s

        # Handle glued plus at end: '4.5+' or '4.5 +' -> '4.5'
        try:
            import re

            s2 = re.sub(r"\s*\+\s*$", "", s)
            return s2.strip()
        except Exception:
            return s

    def _header_text(col_index: int) -> str:
        try:
            hi = table.horizontalHeaderItem(int(col_index))
            return "" if hi is None else str(hi.text() or "").strip()
        except Exception:
            return ""

    header_by_table_col = {int(c): _header_text(int(c)) for c in cols}
    header_lower_to_table_col = {
        str(v or "").strip().lower(): int(k)
        for k, v in header_by_table_col.items()
        if str(v or "").strip()
    }

    col_department = header_lower_to_table_col.get("phòng ban")
    col_title_name = header_lower_to_table_col.get("chức vụ")

    col_emp_code = header_lower_to_table_col.get("mã nv")
    col_full_name = header_lower_to_table_col.get("tên nhân viên")
    col_date = header_lower_to_table_col.get("ngày")
    col_weekday = header_lower_to_table_col.get("thứ")

    col_in1 = header_lower_to_table_col.get("vào 1")
    col_out1 = header_lower_to_table_col.get("ra 1")
    col_in2 = header_lower_to_table_col.get("vào 2")
    col_out2 = header_lower_to_table_col.get("ra 2")
    col_in3 = header_lower_to_table_col.get("vào 3")
    col_out3 = header_lower_to_table_col.get("ra 3")

    col_work = header_lower_to_table_col.get("công")
    col_work_plus = header_lower_to_table_col.get("công +")
    col_hours = header_lower_to_table_col.get("giờ")
    col_hours_plus = header_lower_to_table_col.get("giờ +")
    col_late = header_lower_to_table_col.get("trễ")
    col_early = header_lower_to_table_col.get("sớm")
    col_tc1 = header_lower_to_table_col.get("tc1")
    col_tc2 = header_lower_to_table_col.get("tc2")
    col_tc3 = header_lower_to_table_col.get("tc3")
    col_leave = header_lower_to_table_col.get("kh")

    try:
        row_count = int(table.rowCount())
    except Exception:
        row_count = 0

    rows_source = list(range(row_count))
    if row_indexes is not None:
        rows_source = [int(r) for r in (row_indexes or []) if 0 <= int(r) < row_count]

    def _collect_unique_texts(col: int | None) -> list[str]:
        if col is None:
            return []
        seen: set[str] = set()
        out: list[str] = []
        for rr in rows_source:
            try:
                it = table.item(int(rr), int(col))
                txt = "" if it is None else str(it.text() or "").strip()
            except Exception:
                txt = ""
            if not txt:
                continue
            if txt in seen:
                continue
            seen.add(txt)
            out.append(txt)
        return out

    def _fmt_list(items: list[str]) -> str:
        items = [str(x or "").strip() for x in (items or []) if str(x or "").strip()]
        if not items:
            return ""
        return ", ".join(items)

    dept_txt = _fmt_list(_collect_unique_texts(col_department))
    title_txt = _fmt_list(_collect_unique_texts(col_title_name))
    if not str(dept_txt or "").strip() and department_text is not None:
        dept_txt = str(department_text or "").strip()
    if not str(title_txt or "").strip() and title_text is not None:
        title_txt = str(title_text or "").strip()

    # Decide whether we can export the monthly template (same month only)
    same_month = bool(
        from_d
        and to_d
        and from_d <= to_d
        and from_d.year == to_d.year
        and from_d.month == to_d.month
    )

    can_monthly = bool(
        same_month and col_emp_code is not None and col_full_name is not None
    )

    # Build monthly template only when we have a single-month range and required columns.
    table_ranges: list[tuple[int, int]] = []
    if can_monthly:
        assert from_d is not None and to_d is not None

        # Days columns: only the selected date range within the month.
        days_in_month = int(monthrange(to_d.year, to_d.month)[1])
        start_day = int(from_d.day)
        end_day = int(to_d.day)
        start_day = max(1, min(start_day, days_in_month))
        end_day = max(1, min(end_day, days_in_month))
        if start_day > end_day:
            start_day, end_day = end_day, start_day
        day_list = list(range(int(start_day), int(end_day) + 1))
        day_count = int(len(day_list))

        # Template columns count: 4 fixed + day columns + 16 summary columns
        # Summary columns layout:
        # Ngày công(2), Giờ công(2), Vào trễ(2), Ra sớm(2), Tăng ca(3), Vắng KP(1), Ngày nghỉ(4)
        summary_cols = 16
        total_cols = 4 + day_count + summary_cols
        grid_ncols = int(total_cols)

        # Re-map worksheet-wide merged rows to new width
        def _merge_row_full(r: int) -> None:
            if total_cols <= 1:
                return
            ws.merge_cells(
                start_row=r, start_column=1, end_row=r, end_column=total_cols
            )

        # Rebuild header rows 1..5 to match new width
        # (they were previously merged using old ncols)
        ws.unmerge_cells(str(ws.merged_cells)) if False else None  # no-op placeholder

        # We already wrote company header using _merge_full_row which used initial ncols.
        # For correctness, re-merge rows 1..5 to the new width.
        try:
            for r in range(1, 6):
                # safe unmerge if previously merged
                for m in list(ws.merged_cells.ranges):
                    if m.min_row == r and m.max_row == r:
                        ws.unmerge_cells(str(m))
                _merge_row_full(r)
        except Exception:
            pass

        # Build table header rows 6-7 according to template
        row6 = 6
        row7 = 7

        # A..D headers
        ws.cell(row=row6, column=1, value="STT").font = header_font
        ws.cell(row=row6, column=2, value="Mã nhân viên").font = header_font
        ws.cell(row=row6, column=3, value="Tên nhân viên").font = header_font
        ws.cell(row=row6, column=4, value="").font = header_font

        # Merge vertical for A..D
        for c in (1, 2, 3, 4):
            ws.merge_cells(start_row=row6, start_column=c, end_row=row7, end_column=c)
            ws.cell(row=row6, column=c).alignment = grid_center

        # Day headers: start at column 5 (E)
        day_start_col = 5
        for i, day in enumerate(day_list):
            col = day_start_col + int(i)
            ws.cell(row=row6, column=col, value=str(day)).font = header_font
            ws.cell(row=row6, column=col).alignment = grid_center
            d_obj = date(to_d.year, to_d.month, day)
            ws.cell(
                row=row7,
                column=col,
                value=_vn_weekday(d_obj).replace("Thứ ", "T").replace("Chủ nhật", "CN"),
            ).alignment = grid_center

        # Summary columns start
        s0 = day_start_col + day_count
        # Ngày công
        ws.cell(row=row6, column=s0, value="Ngày công").font = header_font
        ws.merge_cells(start_row=row6, start_column=s0, end_row=row6, end_column=s0 + 1)
        ws.cell(row=row7, column=s0, value="NT").alignment = grid_center
        ws.cell(row=row7, column=s0 + 1, value="CT").alignment = grid_center
        # Giờ công
        s1 = s0 + 2
        ws.cell(row=row6, column=s1, value="Giờ công").font = header_font
        ws.merge_cells(start_row=row6, start_column=s1, end_row=row6, end_column=s1 + 1)
        ws.cell(row=row7, column=s1, value="NT").alignment = grid_center
        ws.cell(row=row7, column=s1 + 1, value="CT").alignment = grid_center
        # Vào trễ
        s2 = s1 + 2
        ws.cell(row=row6, column=s2, value="Vào trễ").font = header_font
        ws.merge_cells(start_row=row6, start_column=s2, end_row=row6, end_column=s2 + 1)
        ws.cell(row=row7, column=s2, value="Lần").alignment = grid_center
        ws.cell(row=row7, column=s2 + 1, value="Phút").alignment = grid_center
        # Ra sớm
        s3 = s2 + 2
        ws.cell(row=row6, column=s3, value="Ra sớm").font = header_font
        ws.merge_cells(start_row=row6, start_column=s3, end_row=row6, end_column=s3 + 1)
        ws.cell(row=row7, column=s3, value="Lần").alignment = grid_center
        ws.cell(row=row7, column=s3 + 1, value="Phút").alignment = grid_center
        # Tăng ca (giờ)
        s4 = s3 + 2
        ws.cell(row=row6, column=s4, value="Tăng ca (giờ)").font = header_font
        ws.merge_cells(start_row=row6, start_column=s4, end_row=row6, end_column=s4 + 2)
        ws.cell(row=row7, column=s4, value="TC1").alignment = grid_center
        ws.cell(row=row7, column=s4 + 1, value="TC2").alignment = grid_center
        ws.cell(row=row7, column=s4 + 2, value="TC3").alignment = grid_center
        # Vắng KP
        s5 = s4 + 3
        ws.cell(row=row6, column=s5, value="Vắng KP").font = header_font
        ws.merge_cells(start_row=row6, start_column=s5, end_row=row7, end_column=s5)
        ws.cell(row=row6, column=s5).alignment = grid_center
        # Ngày nghỉ
        s6 = s5 + 1
        ws.cell(row=row6, column=s6, value="Ngày nghỉ").font = header_font
        ws.merge_cells(start_row=row6, start_column=s6, end_row=row6, end_column=s6 + 3)
        ws.cell(row=row7, column=s6, value="OM").alignment = grid_center
        ws.cell(row=row7, column=s6 + 1, value="TS").alignment = grid_center
        ws.cell(row=row7, column=s6 + 2, value="R").alignment = grid_center
        ws.cell(row=row7, column=s6 + 3, value="Le").alignment = grid_center

        # Align row6/7 summary cells center
        for c in range(s0, total_cols + 1):
            ws.cell(row=row6, column=c).alignment = grid_center
            ws.cell(row=row7, column=c).alignment = grid_center

        _merge_row_full(8)
        ws.cell(
            row=8,
            column=1,
            value=(f"Phòng ban: {dept_txt}" if dept_txt else "Phòng ban:"),
        ).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        try:
            ws.row_dimensions[8].height = max(30, int(ws.row_dimensions[8].height or 0))
        except Exception:
            pass

        _merge_row_full(9)
        ws.cell(
            row=9,
            column=1,
            value=(f"Chức vụ: {title_txt}" if title_txt else "Chức vụ:"),
        ).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        try:
            ws.row_dimensions[9].height = max(30, int(ws.row_dimensions[9].height or 0))
        except Exception:
            pass

        # Collect source rows
        def _cell_text(r: int, c: int | None) -> str:
            if c is None:
                return ""
            it = table.item(int(r), int(c))
            return "" if it is None else str(it.text() or "").strip()

        def _is_col_visible(c: int | None) -> bool:
            if c is None:
                return False
            try:
                return not bool(table.isColumnHidden(int(c)))
            except Exception:
                return True

        def _first_number(s: str) -> str:
            # Extract first numeric token from strings like "1 X" or "2.5 +".
            import re

            t = str(s or "").strip()
            m = re.search(r"[-+]?\d+(?:[\.,]\d+)?", t)
            return "" if m is None else m.group(0).replace(",", ".")

        def _to_decimal(s: str) -> "Decimal":
            from decimal import Decimal

            t = _first_number(s)
            if not t:
                return Decimal("0")
            try:
                return Decimal(t)
            except Exception:
                return Decimal("0")

        def _to_int_minutes(s: str) -> int:
            t = _first_number(s)
            if not t:
                return 0
            try:
                return int(float(t))
            except Exception:
                return 0

        def _fmt_decimal(d) -> str:
            # Pretty format for Excel: drop trailing zeros.
            try:
                from decimal import Decimal

                if not isinstance(d, Decimal):
                    d = Decimal(str(d))
                s = format(d.normalize(), "f")
                if s == "-0":
                    s = "0"
                return s
            except Exception:
                try:
                    f = float(d)
                    if abs(f - int(f)) < 1e-9:
                        return str(int(f))
                    return str(f)
                except Exception:
                    return "0"

        employees: list[tuple[str, str]] = []
        by_emp_day: dict[tuple[str, str], dict[int, dict[str, str]]] = {}

        for r in rows_source:
            code = _cell_text(int(r), col_emp_code)
            name = _cell_text(int(r), col_full_name)
            key = (code, name)
            if key not in by_emp_day:
                by_emp_day[key] = {}
                employees.append(key)

            d_obj = (
                _parse_date_any(_cell_text(int(r), col_date))
                if col_date is not None
                else None
            )
            if d_obj is None or d_obj.year != to_d.year or d_obj.month != to_d.month:
                continue
            day = int(d_obj.day)
            if day < int(start_day) or day > int(end_day):
                continue

            by_emp_day[key].setdefault(day, {})
            rec = by_emp_day[key][day]
            rec["in1"] = _cell_text(int(r), col_in1)
            rec["out1"] = _cell_text(int(r), col_out1)
            rec["in2"] = _cell_text(int(r), col_in2)
            rec["out2"] = _cell_text(int(r), col_out2)
            rec["in3"] = _cell_text(int(r), col_in3)
            rec["out3"] = _cell_text(int(r), col_out3)

            # Summary-related values (strings as displayed in UI table)
            rec["work"] = _cell_text(int(r), col_work)
            rec["hours"] = _cell_text(int(r), col_hours)
            rec["late"] = _cell_text(int(r), col_late)
            rec["early"] = _cell_text(int(r), col_early)
            rec["tc1"] = _cell_text(int(r), col_tc1)
            rec["tc2"] = _cell_text(int(r), col_tc2)
            rec["tc3"] = _cell_text(int(r), col_tc3)
            rec["leave"] = _cell_text(int(r), col_leave)

        # Determine in/out lines for each employee.
        # User requirement: only export what is shown on the table.
        # So hidden columns should not appear as extra rows in the template.
        # We also respect the user's selected 2/4/6 cap via force_exclude_headers.
        excluded_headers = {
            str(x or "").strip() for x in (force_exclude_headers or set())
        }

        excluded_headers_norm = {h.lower() for h in excluded_headers if str(h).strip()}

        # User-selected cap via force_exclude_headers:
        # - exclude pair2 => max_pairs=1
        # - exclude pair3 => max_pairs=2
        max_pairs = 3

        # First: if columns are not present or hidden, reduce pairs.
        if not (_is_col_visible(col_in2) or _is_col_visible(col_out2)):
            max_pairs = 1
        elif not (_is_col_visible(col_in3) or _is_col_visible(col_out3)):
            max_pairs = 2

        # Then: explicit excludes from settings (handle common header variants).
        if {"vào 2", "vao 2", "ra 2"} & excluded_headers_norm:
            max_pairs = 1
        elif {"vào 3", "vao 3", "ra 3"} & excluded_headers_norm:
            max_pairs = min(max_pairs, 2)

        def _need_pair2(_emp_code: str, _key: tuple[str, str]) -> bool:
            return bool(max_pairs >= 2)

        def _need_pair3(_emp_code: str, _key: tuple[str, str]) -> bool:
            return bool(max_pairs >= 3)

        # Data starts at row 10
        cur = 10
        employee_blocks: list[tuple[int, int]] = []  # (start_row, block_height)
        for idx, (code, name) in enumerate(employees, start=1):
            key = (code, name)
            pair2 = _need_pair2(code, key)
            pair3 = _need_pair3(code, key)
            lines = [("Vào1", "in1"), ("Ra1", "out1")]
            if pair2:
                lines += [("Vào2", "in2"), ("Ra2", "out2")]
            if pair3:
                lines += [("Vào3", "in3"), ("Ra3", "out3")]

            block_h = len(lines)
            employee_blocks.append((int(cur), int(block_h)))

            # Merge A..C vertically for employee block
            for col in (1, 2, 3):
                ws.merge_cells(
                    start_row=cur,
                    start_column=col,
                    end_row=cur + block_h - 1,
                    end_column=col,
                )

            ws.cell(row=cur, column=1, value=idx).alignment = grid_center
            ws.cell(row=cur, column=2, value=code).alignment = grid_center
            ws.cell(row=cur, column=3, value=name).alignment = grid_center

            # Summary values: use first available day record
            days = by_emp_day.get(key, {})
            sample_day = next(iter(sorted(days.keys())), None)
            sample = days.get(sample_day, {}) if sample_day is not None else {}

            # Write line labels + day cells
            for line_i, (label, field) in enumerate(lines):
                rr = cur + line_i
                ws.cell(row=rr, column=4, value=label).alignment = grid_center
                for i, day in enumerate(day_list):
                    col = day_start_col + int(i)
                    rec = days.get(day, {})
                    ws.cell(
                        row=rr,
                        column=col,
                        value=_strip_export_symbols(rec.get(field, "") or "", ""),
                    ).alignment = grid_center

            # Merge status symbols V/OFF/Lễ when present in IN cells into their IN/OUT pair.
            # In monthly template, IN/OUT are two separate rows => merge vertically for that day.
            try:
                for i, day in enumerate(day_list):
                    col = day_start_col + int(i)
                    rec = days.get(day, {})

                    def _merge_pair_rows(
                        in_row: int, out_row: int, in_key: str
                    ) -> None:
                        in_val = rec.get(in_key, "")
                        if not _is_merge_status_symbol(in_val):
                            return
                        try:
                            ws.merge_cells(
                                start_row=int(in_row),
                                start_column=int(col),
                                end_row=int(out_row),
                                end_column=int(col),
                            )
                        except Exception:
                            return
                        # Put symbol on the top cell only.
                        try:
                            ws.cell(
                                row=int(in_row),
                                column=int(col),
                                value=_norm_symbol_text(in_val),
                            ).alignment = grid_center
                            ws.cell(
                                row=int(out_row), column=int(col), value=""
                            ).alignment = grid_center
                        except Exception:
                            pass

                    # Pair 1 is always present
                    _merge_pair_rows(cur + 0, cur + 1, "in1")
                    # Pair 2 and 3 depend on export cap
                    if pair2:
                        base = 2
                        _merge_pair_rows(cur + base, cur + base + 1, "in2")
                    if pair3:
                        base = 4 if pair2 else 2
                        _merge_pair_rows(cur + base, cur + base + 1, "in3")
            except Exception:
                pass

            # Merge summary columns vertically and set values on first row
            def _merge_vert(c: int) -> None:
                ws.merge_cells(
                    start_row=cur,
                    start_column=c,
                    end_row=cur + block_h - 1,
                    end_column=c,
                )

            for c in range(s0, total_cols + 1):
                _merge_vert(c)
                ws.cell(row=cur, column=c, value="0").alignment = grid_center

            # Compute summaries per the requested rules.
            # - Ngày công NT: Mon..Sat full-work days (sum integer work when work is full-day)
            # - Ngày công CT: Sunday full-work days
            # - Giờ công NT: total hours of non-full days (Mon..Sat)
            # - Giờ công CT: total hours of non-full days (Sunday)
            # - Late/Early: count (days) + total minutes
            # - TC1/2/3: total overtime hours from tc columns
            # - Lễ: total holiday days
            from decimal import Decimal, ROUND_DOWN

            full_nt = 0
            full_ct = 0
            hours_nt = Decimal("0")
            hours_ct = Decimal("0")
            late_times = 0
            late_minutes = 0
            early_times = 0
            early_minutes = 0
            tc1_sum = Decimal("0")
            tc2_sum = Decimal("0")
            tc3_sum = Decimal("0")
            holiday_days = 0

            absent_v_times = 0

            def _is_absent_v(rec0: dict[str, str]) -> bool:
                # Any in/out value equals absent symbol text (V)
                for k in ("in1", "out1", "in2", "out2", "in3", "out3"):
                    v0 = _norm_symbol_text(rec0.get(k, ""))
                    if v0.lower() == "v":
                        return True
                return False

            def _is_holiday_day(rec0: dict[str, str]) -> bool:
                # Heuristic: any in/out value equals holiday symbol text (Le/Lễ).
                for k in ("in1", "out1", "in2", "out2", "in3", "out3"):
                    v0 = _norm_symbol_text(rec0.get(k, ""))
                    if v0.lower() in {"le", "lễ"}:
                        return True
                return False

            for day in day_list:
                d_obj = date(to_d.year, to_d.month, int(day))
                is_sunday = int(d_obj.weekday()) == 6
                rec0 = days.get(int(day), {})
                if not rec0:
                    continue

                if _is_absent_v(rec0):
                    absent_v_times += 1

                work_d = _to_decimal(str(rec0.get("work", "") or ""))
                hours_d = _to_decimal(str(rec0.get("hours", "") or ""))
                late_m = _to_int_minutes(str(rec0.get("late", "") or ""))
                early_m = _to_int_minutes(str(rec0.get("early", "") or ""))
                tc1_d = _to_decimal(str(rec0.get("tc1", "") or ""))
                tc2_d = _to_decimal(str(rec0.get("tc2", "") or ""))
                tc3_d = _to_decimal(str(rec0.get("tc3", "") or ""))

                if late_m > 0:
                    late_times += 1
                    late_minutes += int(late_m)
                if early_m > 0:
                    early_times += 1
                    early_minutes += int(early_m)

                if tc1_d > 0:
                    tc1_sum += tc1_d
                if tc2_d > 0:
                    tc2_sum += tc2_d
                if tc3_d > 0:
                    tc3_sum += tc3_d

                if _is_holiday_day(rec0):
                    holiday_days += 1

                # Full day check: treat integer part >= 1 and no fractional as full.
                int_part = work_d.quantize(Decimal("1"), rounding=ROUND_DOWN)
                is_full = bool(work_d >= Decimal("1") and work_d == int_part)

                if is_full:
                    if is_sunday:
                        full_ct += int(int_part)
                    else:
                        full_nt += int(int_part)
                else:
                    if hours_d > 0:
                        if is_sunday:
                            hours_ct += hours_d
                        else:
                            hours_nt += hours_d

            # Write summaries to first row (merged cells)
            # Requirement: when exporting Excel, do not export symbol 'X'.
            ws.cell(
                row=cur,
                column=s0,
                value=(str(int(full_nt)) if int(full_nt) > 0 else "0"),
            ).alignment = grid_center
            ws.cell(
                row=cur,
                column=s0 + 1,
                value=(str(int(full_ct)) if int(full_ct) > 0 else "0"),
            ).alignment = grid_center

            ws.cell(row=cur, column=s1, value=_fmt_decimal(hours_nt)).alignment = (
                grid_center
            )
            ws.cell(row=cur, column=s1 + 1, value=_fmt_decimal(hours_ct)).alignment = (
                grid_center
            )

            ws.cell(row=cur, column=s2, value=str(late_times)).alignment = grid_center
            ws.cell(row=cur, column=s2 + 1, value=str(late_minutes)).alignment = (
                grid_center
            )
            ws.cell(row=cur, column=s3, value=str(early_times)).alignment = grid_center
            ws.cell(row=cur, column=s3 + 1, value=str(early_minutes)).alignment = (
                grid_center
            )

            ws.cell(row=cur, column=s4, value=_fmt_decimal(tc1_sum)).alignment = (
                grid_center
            )
            ws.cell(row=cur, column=s4 + 1, value=_fmt_decimal(tc2_sum)).alignment = (
                grid_center
            )
            ws.cell(row=cur, column=s4 + 2, value=_fmt_decimal(tc3_sum)).alignment = (
                grid_center
            )

            # Vắng KP: count days that contain absent symbol 'V'
            try:
                ws.cell(
                    row=cur, column=s5, value=str(int(absent_v_times))
                ).alignment = grid_center
            except Exception:
                ws.cell(row=cur, column=s5, value="0").alignment = grid_center

            # Ngày nghỉ (Le) only
            try:
                ws.cell(
                    row=cur, column=s6 + 3, value=str(int(holiday_days))
                ).alignment = grid_center
            except Exception:
                pass

            # Next employee block
            end_block = cur + block_h - 1
            cur = end_block + 1

        # One full table range (header + meta rows + all employees)
        last_table_row = max(9, cur - 1)
        last_table_row = max(9, cur - 1)
        table_ranges.append((row6, last_table_row))

    else:
        # Fallback: original behavior (export visible table rows)
        header_row = 6
        for excel_col, table_col in enumerate(cols, start=1):
            label = header_by_table_col.get(int(table_col), "")
            cell = ws.cell(row=header_row, column=excel_col, value=label)
            cell.font = header_font
            cell.alignment = center

        # Rows 8-9: merged labels (include values based on exported rows)
        try:
            _merge_full_row(8)
            ws.cell(
                row=8,
                column=1,
                value=(f"Phòng ban: {dept_txt}" if dept_txt else "Phòng ban:"),
            ).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.row_dimensions[8].height = max(30, int(ws.row_dimensions[8].height or 0))
        except Exception:
            pass

        try:
            _merge_full_row(9)
            ws.cell(
                row=9,
                column=1,
                value=(f"Chức vụ: {title_txt}" if title_txt else "Chức vụ:"),
            ).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.row_dimensions[9].height = max(30, int(ws.row_dimensions[9].height or 0))
        except Exception:
            pass

        # Data starts at row 10 to keep row 8/9 reserved for meta.
        start_row = 10
        excel_col_by_table_col = {
            int(tc): int(ec) for ec, tc in enumerate(cols, start=1)
        }

        def _try_merge_pair_cols(
            excel_row: int, in_table_col: int | None, out_table_col: int | None
        ) -> None:
            if in_table_col is None or out_table_col is None:
                return
            if (
                int(in_table_col) not in excel_col_by_table_col
                or int(out_table_col) not in excel_col_by_table_col
            ):
                return
            in_excel_col = int(excel_col_by_table_col[int(in_table_col)])
            out_excel_col = int(excel_col_by_table_col[int(out_table_col)])
            if in_excel_col == out_excel_col:
                return
            try:
                in_cell = ws.cell(row=int(excel_row), column=int(in_excel_col))
                in_txt = _norm_symbol_text(in_cell.value)
            except Exception:
                return
            if not _is_merge_status_symbol(in_txt):
                return
            # Ensure left-to-right merge
            c1, c2 = (
                (in_excel_col, out_excel_col)
                if in_excel_col < out_excel_col
                else (out_excel_col, in_excel_col)
            )
            try:
                # Clear the OUT cell (or the other side) to avoid duplicate text.
                ws.cell(row=int(excel_row), column=int(c2), value="")
                ws.merge_cells(
                    start_row=int(excel_row),
                    start_column=int(c1),
                    end_row=int(excel_row),
                    end_column=int(c2),
                )
                ws.cell(row=int(excel_row), column=int(c1), value=in_txt).alignment = (
                    left
                )
            except Exception:
                pass

        for out_idx, r in enumerate(rows_source):
            excel_row = start_row + int(out_idx)
            for excel_col, table_col in enumerate(cols, start=1):
                try:
                    item = table.item(int(r), int(table_col))
                    txt = "" if item is None else str(item.text() or "")
                except Exception:
                    txt = ""
                header_label = header_by_table_col.get(int(table_col), "")
                cell = ws.cell(
                    row=excel_row,
                    column=excel_col,
                    value=_strip_export_symbols(txt, header_label),
                )
                cell.alignment = left

            # Apply merge rule for detail export: if IN cell is V/OFF/Lễ then merge IN/OUT pair.
            _try_merge_pair_cols(excel_row, col_in1, col_out1)
            _try_merge_pair_cols(excel_row, col_in2, col_out2)
            _try_merge_pair_cols(excel_row, col_in3, col_out3)

        last_data_row = (start_row - 1) + len(rows_source)
        table_ranges.append((header_row, last_data_row))

    # Borders
    thin = Side(style="thin")
    dotted = Side(style="dotted")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r0, r1 in table_ranges:
        if r0 <= 0 or r1 <= 0:
            continue
        for rr in range(int(r0), int(r1) + 1):
            for cc in range(1, int(grid_ncols) + 1):
                try:
                    ws.cell(row=int(rr), column=int(cc)).border = grid_border
                except Exception:
                    pass

    # Dotted separators between in/out rows (match template look):
    # Apply dotted horizontal borders within each employee block, starting from column 4 (label) onward.
    if can_monthly:
        try:

            def _replace_border(
                b: Border, *, top: Side | None = None, bottom: Side | None = None
            ) -> Border:
                return Border(
                    left=b.left,
                    right=b.right,
                    top=(top if top is not None else b.top),
                    bottom=(bottom if bottom is not None else b.bottom),
                    diagonal=b.diagonal,
                    diagonal_direction=b.diagonal_direction,
                    outline=b.outline,
                    vertical=b.vertical,
                    horizontal=b.horizontal,
                )

            for start_row, block_h in employee_blocks:
                start_row = int(start_row)
                block_h = int(block_h)
                if block_h <= 1:
                    continue
                for boundary_row in range(start_row, start_row + block_h - 1):
                    r_top = int(boundary_row)
                    r_bottom = int(boundary_row + 1)
                    for cc in range(4, int(grid_ncols) + 1):
                        c1 = ws.cell(row=r_top, column=int(cc))
                        c2 = ws.cell(row=r_bottom, column=int(cc))
                        c1.border = _replace_border(c1.border, bottom=dotted)
                        c2.border = _replace_border(c2.border, top=dotted)
        except Exception:
            pass

    # Column widths
    if can_monthly:
        # Template-like widths: make day columns wide enough so values don't wrap.
        # (This ensures column W and other day columns have enough width.)
        ws.column_dimensions[get_column_letter(1)].width = 4.6
        ws.column_dimensions[get_column_letter(2)].width = 13.3
        ws.column_dimensions[get_column_letter(3)].width = 20.0
        ws.column_dimensions[get_column_letter(4)].width = 4.8

        # Day + summary columns: use a wider default width to avoid line breaks.
        for c in range(5, int(grid_ncols) + 1):
            ws.column_dimensions[get_column_letter(int(c))].width = 13.0
    else:
        # Fallback: widths from QTableWidget
        for excel_col, table_col in enumerate(cols, start=1):
            try:
                px = int(table.columnWidth(int(table_col)))
            except Exception:
                px = 120
            width = max(10, min(60, int(round(px / 7))))
            ws.column_dimensions[get_column_letter(excel_col)].width = width

    # Basic row heights (will be adjusted later)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[6].height = 20

    # Append creator + note
    creator_txt = str(creator or "").strip()
    note_txt = str(note_text or "").strip()
    if creator_txt or note_txt:
        base_row = int(ws.max_row or 0) + 2

        def _merge_row(r: int) -> None:
            if int(grid_ncols) <= 1:
                return
            ws.merge_cells(
                start_row=r,
                start_column=1,
                end_row=r,
                end_column=int(grid_ncols),
            )

        if creator_txt:
            _merge_row(base_row)
            f, a = _norm_style(creator_style, default_align="left")
            c = ws.cell(
                row=base_row, column=1, value=f"Người tạo danh sách: {creator_txt}"
            )
            c.font = f
            c.alignment = a
            base_row += 1

        if note_txt:
            _merge_row(base_row)
            f, a = _norm_style(note_style, default_align="left")
            # Note is typically multi-line
            a = Alignment(
                horizontal=a.horizontal,
                vertical="top",
                wrap_text=True,
            )
            ncell = ws.cell(row=base_row, column=1, value=f"Ghi chú: {note_txt}")
            ncell.font = f
            ncell.alignment = a
            try:
                ws.row_dimensions[base_row].height = 40
            except Exception:
                pass

    # Content rows: fixed height 15 (apply to data only)
    try:
        max_row = int(ws.max_row or 0)
        data_start = 10
        data_end = max_row
        try:
            if table_ranges:
                # table_ranges includes header + data; keep only the last row
                data_end = int(max(r1 for _r0, r1 in table_ranges))
        except Exception:
            data_end = max_row

        for r in range(int(data_start), int(data_end) + 1):
            h = ws.row_dimensions[int(r)].height
            # Preserve explicit note height (usually 40)
            if h is not None and float(h) >= 39:
                continue
            ws.row_dimensions[int(r)].height = 15
    except Exception:
        pass

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except PermissionError:
        return (
            False,
            f"Không có quyền ghi vào thư mục: {path.parent}.",
        )
    except Exception:
        # Keep generic error message (caller will display).
        return False, "Không thể tạo thư mục để lưu file xuất."

    # Freeze panes as requested (keep header + left columns visible)
    # Excel notation: D10 freezes rows 1..9 and columns A..C.
    try:
        ws.freeze_panes = "D10"
    except Exception:
        pass

    try:
        wb.save(str(path))
    except PermissionError:
        # Common on Windows when the file is opened by Excel.
        return (
            False,
            f"File đang được mở hoặc không có quyền ghi đè: {path}. Vui lòng đóng file Excel đang mở rồi xuất lại.",
        )
    except Exception:
        return False, f"Không thể lưu file xuất: {path}"

    # Auto-open exported file (best-effort)
    try:
        import os

        if os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
    except Exception:
        pass

    return True, f"Đã xuất dữ liệu: {path}"
