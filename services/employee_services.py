"""services.employee_services

Service cho màn Thông tin Nhân viên:
- list employees theo filter
- export/import CSV
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any
from typing import Callable

from repository.employee_repository import EmployeeRepository
from services.department_services import DepartmentService
from services.title_services import TitleService


class EmployeeService:
    def __init__(
        self,
        repo: EmployeeRepository | None = None,
        department_service: DepartmentService | None = None,
        title_service: TitleService | None = None,
    ) -> None:
        self._repo = repo or EmployeeRepository()
        self._department_service = department_service or DepartmentService()
        self._title_service = title_service or TitleService()

    def list_departments_tree_rows(self) -> list[tuple[int, int | None, str, str]]:
        models = self._department_service.list_departments()
        return [
            (m.id, m.parent_id, m.department_name, m.department_note) for m in models
        ]

    def list_employees(self, filters: dict) -> list[dict[str, Any]]:
        return self._repo.list_employees(
            employee_code=str(filters.get("employee_code") or "").strip() or None,
            full_name=str(filters.get("full_name") or "").strip() or None,
            department_id=filters.get("department_id"),
        )

    def list_departments_dropdown(self) -> list[tuple[int, str]]:
        models = self._department_service.list_departments()
        items: list[tuple[int, str]] = []
        for m in models:
            try:
                items.append((int(m.id), str(m.department_name)))
            except Exception:
                continue
        items.sort(key=lambda x: x[1].lower())
        return items

    def list_titles_dropdown(self) -> list[tuple[int, str]]:
        models = self._title_service.list_titles()
        items: list[tuple[int, str]] = []
        for m in models:
            try:
                items.append((int(m.id), str(m.title_name)))
            except Exception:
                continue
        items.sort(key=lambda x: x[1].lower())
        return items

    def list_issue_places_dropdown(self) -> list[str]:
        try:
            return self._repo.list_distinct_id_issue_places()
        except Exception:
            return []

    def get_employee(self, employee_id: int) -> dict[str, Any] | None:
        row = self._repo.get_employee(int(employee_id))
        if not row:
            return None

        def to_str(v: Any) -> Any:
            if v is None:
                return None
            try:
                return v.isoformat()
            except Exception:
                return v

        # Normalize dates to ISO strings for UI
        for k in (
            "start_date",
            "date_of_birth",
            "id_issue_date",
            "contract1_sign_date",
            "contract1_expire_date",
            "contract2_sign_date",
            "child_dob_1",
            "child_dob_2",
            "child_dob_3",
            "child_dob_4",
        ):
            row[k] = to_str(row.get(k))
        return row

    def export_csv(self, file_path: str, filters: dict) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file xuất."
        if path.suffix.lower() != ".csv":
            path = path.with_suffix(".csv")

        rows = self.list_employees(filters)

        headers = [
            "id",
            "employee_code",
            "full_name",
            "start_date",
            "title_name",
            "department_name",
            "date_of_birth",
            "gender",
            "national_id",
            "id_issue_date",
            "id_issue_place",
            "address",
            "phone",
            "insurance_no",
            "tax_code",
            "degree",
            "major",
            "contract1_signed",
            "contract1_no",
            "contract1_sign_date",
            "contract1_expire_date",
            "contract2_indefinite",
            "contract2_no",
            "contract2_sign_date",
            "children_count",
            "child_dob_1",
            "child_dob_2",
            "child_dob_3",
            "child_dob_4",
            "note",
        ]

        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in headers})

        return True, f"Đã xuất {len(rows)} dòng: {path}"

    def export_xlsx(self, file_path: str, filters: dict) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file xuất."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện xuất Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        rows = self.list_employees(filters)

        columns: list[tuple[str, str]] = [
            ("STT", "stt"),
            ("Mã NV", "employee_code"),
            ("Họ và tên", "full_name"),
            ("Ngày vào làm", "start_date"),
            ("Chức Vụ", "title_name"),
            ("Phòng Ban", "department_name"),
            ("Ngày tháng năm sinh", "date_of_birth"),
            ("Giới tính", "gender"),
            ("CCCD/CMT", "national_id"),
            ("Ngày Cấp", "id_issue_date"),
            ("Nơi Cấp", "id_issue_place"),
            ("Địa chỉ", "address"),
            ("Số điện thoại", "phone"),
            ("Số Bảo Hiểm", "insurance_no"),
            ("Mã số Thuế TNCN", "tax_code"),
            ("Bằng cấp", "degree"),
            ("Chuyên ngành", "major"),
            ("HĐLĐ (ký lần 1)", "contract1_signed"),
            ("Số HĐLĐ (lần 1)", "contract1_no"),
            ("Ngày ký (lần 1)", "contract1_sign_date"),
            ("Ngày hết hạn (lần 1)", "contract1_expire_date"),
            ("HĐLĐ ký không thời hạn", "contract2_indefinite"),
            ("Số HĐLĐ (không thời hạn)", "contract2_no"),
            ("Ngày ký (không thời hạn)", "contract2_sign_date"),
            ("Số con", "children_count"),
            ("Ngày sinh con 1", "child_dob_1"),
            ("Ngày sinh con 2", "child_dob_2"),
            ("Ngày sinh con 3", "child_dob_3"),
            ("Ngày sinh con 4", "child_dob_4"),
            ("Ghi chú", "note"),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        header_font = Font(bold=True)
        for col_idx, (label, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font

        for row_idx, r in enumerate(rows, start=2):
            for col_idx, (_label, key) in enumerate(columns, start=1):
                v = r.get(key)
                if key in {"contract1_signed", "contract2_indefinite"}:
                    v = "1" if bool(v) else "0"
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=("" if v is None else v),
                )

        # basic width
        for col_idx, (label, _key) in enumerate(columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                max(12, min(40, len(label) + 6))
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return True, f"Đã xuất {len(rows)} dòng: {path}"

    def export_employee_template_xlsx(self, file_path: str) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn file mẫu."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện đọc/ghi Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        # Use Vietnamese headers (same spirit as export_xlsx) so users can fill in easily.
        columns: list[tuple[str, str]] = [
            ("STT", "stt"),
            ("MÃ NV", "employee_code"),
            ("HỌ VÀ TÊN", "full_name"),
            ("Ngày vào làm", "start_date"),
            ("Chức Vụ", "title_name"),
            ("Phòng Ban", "department_name"),
            ("Ngày tháng năm sinh", "date_of_birth"),
            ("Giới tính", "gender"),
            ("CCCD/CMT", "national_id"),
            ("Ngày Cấp", "id_issue_date"),
            ("Nơi Cấp", "id_issue_place"),
            ("Địa chỉ", "address"),
            ("Số điện thoại", "phone"),
            ("Số Bảo Hiểm", "insurance_no"),
            ("Mã số Thuế TNCN", "tax_code"),
            ("Bằng cấp", "degree"),
            ("Chuyên ngành", "major"),
            ("HĐLĐ (ký lần 1)", "contract1_signed"),
            ("Số HĐLĐ (lần 1)", "contract1_no"),
            ("Ngày ký (lần 1)", "contract1_sign_date"),
            ("Ngày hết hạn (lần 1)", "contract1_expire_date"),
            ("HĐLĐ ký không thời hạn", "contract2_indefinite"),
            ("Số HĐLĐ (không thời hạn)", "contract2_no"),
            ("Ngày ký (không thời hạn)", "contract2_sign_date"),
            ("Số con", "children_count"),
            ("Ngày sinh con 1", "child_dob_1"),
            ("Ngày sinh con 2", "child_dob_2"),
            ("Ngày sinh con 3", "child_dob_3"),
            ("Ngày sinh con 4", "child_dob_4"),
            ("Ghi chú", "note"),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        header_font = Font(bold=True)
        for col_idx, (label, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font

        for col_idx, (label, _key) in enumerate(columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                max(12, min(40, len(label) + 6))
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return True, f"Đã tạo file mẫu: {path}"

    def read_employees_from_xlsx(
        self, file_path: str
    ) -> tuple[bool, str, list[dict[str, Any]]]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng nhập đường dẫn file Excel.", []
        if not path.exists() or path.suffix.lower() != ".xlsx":
            return False, "Vui lòng chọn file .xlsx hợp lệ.", []

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện đọc Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
                [],
            )

        def norm_header(s: Any) -> str:
            return str(s or "").strip()

        header_to_key: dict[str, str] = {
            # keys
            "employee_code": "employee_code",
            "full_name": "full_name",
            "start_date": "start_date",
            "title_name": "title_name",
            "department_name": "department_name",
            "date_of_birth": "date_of_birth",
            "gender": "gender",
            "national_id": "national_id",
            "id_issue_date": "id_issue_date",
            "id_issue_place": "id_issue_place",
            "address": "address",
            "phone": "phone",
            "insurance_no": "insurance_no",
            "tax_code": "tax_code",
            "degree": "degree",
            "major": "major",
            "contract1_signed": "contract1_signed",
            "contract1_no": "contract1_no",
            "contract1_sign_date": "contract1_sign_date",
            "contract1_expire_date": "contract1_expire_date",
            "contract2_indefinite": "contract2_indefinite",
            "contract2_no": "contract2_no",
            "contract2_sign_date": "contract2_sign_date",
            "children_count": "children_count",
            "child_dob_1": "child_dob_1",
            "child_dob_2": "child_dob_2",
            "child_dob_3": "child_dob_3",
            "child_dob_4": "child_dob_4",
            "note": "note",
            # Vietnamese
            "MÃ NV": "employee_code",
            "Mã NV": "employee_code",
            "HỌ VÀ TÊN": "full_name",
            "Họ và tên": "full_name",
            "Ngày vào làm": "start_date",
            "Chức Vụ": "title_name",
            "Phòng Ban": "department_name",
            "Ngày tháng năm sinh": "date_of_birth",
            "Giới tính": "gender",
            "CCCD/CMT": "national_id",
            "Ngày Cấp": "id_issue_date",
            "Nơi Cấp": "id_issue_place",
            "Địa chỉ": "address",
            "Số điện thoại": "phone",
            "Số Bảo Hiểm": "insurance_no",
            "Mã số Thuế TNCN": "tax_code",
            "Bằng cấp": "degree",
            "Chuyên ngành": "major",
            "HĐLĐ (ký lần 1)": "contract1_signed",
            "Số HĐLĐ (lần 1)": "contract1_no",
            "Ngày ký (lần 1)": "contract1_sign_date",
            "Ngày hết hạn (lần 1)": "contract1_expire_date",
            "HĐLĐ ký không thời hạn": "contract2_indefinite",
            "Số HĐLĐ (không thời hạn)": "contract2_no",
            "Ngày ký (không thời hạn)": "contract2_sign_date",
            "Số con": "children_count",
            "Ngày sinh con 1": "child_dob_1",
            "Ngày sinh con 2": "child_dob_2",
            "Ngày sinh con 3": "child_dob_3",
            "Ngày sinh con 4": "child_dob_4",
            "Ghi chú": "note",
            # ignored
            "STT": "stt",
            "ID": "id",
        }

        def parse_bool(v: Any) -> bool | None:
            if v is None:
                return None
            if isinstance(v, bool):
                return bool(v)
            s = str(v or "").strip().lower()
            if not s:
                return None
            return s in {"1", "true", "yes", "y", "x", "có", "co"}

        def parse_int(v: Any) -> int | None:
            if v is None:
                return None
            s = str(v or "").strip()
            if not s:
                return None
            try:
                return int(float(s))
            except Exception:
                return None

        def parse_date(v: Any) -> str | None:
            if v is None:
                return None
            # openpyxl returns datetime/date objects
            try:
                if hasattr(v, "date"):
                    d = v.date() if hasattr(v, "hour") else v
                    return d.isoformat()
            except Exception:
                pass

            s = str(v or "").strip()
            if not s:
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    return dt.date().isoformat()
                except Exception:
                    continue
            return None

        wb = load_workbook(str(path), data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return False, "File Excel trống.", []

        headers = [norm_header(h) for h in list(header_row or [])]
        col_keys: list[str | None] = []
        for h in headers:
            col_keys.append(header_to_key.get(h))

        out: list[dict[str, Any]] = []
        for r in rows_iter:
            if r is None:
                continue
            item: dict[str, Any] = {}
            empty = True
            for idx, raw in enumerate(list(r)):
                key = col_keys[idx] if idx < len(col_keys) else None
                if not key or key in {"id", "stt"}:
                    continue
                if raw is not None and str(raw).strip() != "":
                    empty = False

                if key in {
                    "start_date",
                    "date_of_birth",
                    "id_issue_date",
                    "contract1_sign_date",
                    "contract1_expire_date",
                    "contract2_sign_date",
                    "child_dob_1",
                    "child_dob_2",
                    "child_dob_3",
                    "child_dob_4",
                }:
                    item[key] = parse_date(raw)
                elif key in {"contract1_signed", "contract2_indefinite"}:
                    item[key] = parse_bool(raw)
                elif key in {"children_count"}:
                    item[key] = parse_int(raw)
                else:
                    s = str(raw or "").strip()
                    item[key] = s if s != "" else None

            if empty:
                continue

            # required fields for UI preview
            code = str(item.get("employee_code") or "").strip()
            name = str(item.get("full_name") or "").strip()
            if code:
                # normalize like create/update
                if code.isdigit():
                    code = code.zfill(5)
                item["employee_code"] = code
            if name:
                item["full_name"] = name

            out.append(item)

        # Add STT for preview
        preview_rows: list[dict[str, Any]] = []
        for idx, it in enumerate(out, start=1):
            row = {"id": None, "stt": idx}
            row.update(it)
            preview_rows.append(row)

        return True, f"Đã đọc {len(preview_rows)} dòng từ Excel.", preview_rows

    def import_employees_rows(
        self,
        rows: list[dict[str, Any]],
        only_new: bool,
        progress_cb: Callable[[int, bool, str, str], None] | None = None,
    ) -> tuple[bool, str]:
        if not rows:
            return False, "Không có dữ liệu để cập nhật."

        # Map title/department name to IDs (best-effort)
        dept_map: dict[str, int] = {}
        for did, dname in self.list_departments_dropdown():
            s = str(dname or "").strip().lower()
            if s:
                dept_map[s] = int(did)

        title_map: dict[str, int] = {}
        for tid, tname in self.list_titles_dropdown():
            s = str(tname or "").strip().lower()
            if s:
                title_map[s] = int(tid)

        def to_bool(v: Any) -> bool:
            return bool(v) if v is not None else False

        def norm_str(v: Any) -> str | None:
            s = str(v or "").strip()
            return s if s else None

        def norm_payload(it: dict[str, Any]) -> dict[str, Any] | None:
            code = str(it.get("employee_code") or "").strip()
            name = str(it.get("full_name") or "").strip()
            if not code or not name:
                return None
            if not code.isdigit():
                return None
            if len(code) > 5:
                return None

            code = code.zfill(5)

            title_name = norm_str(it.get("title_name"))
            dept_name = norm_str(it.get("department_name"))
            title_id = title_map.get(str(title_name or "").lower()) if title_name else None
            dept_id = dept_map.get(str(dept_name or "").lower()) if dept_name else None

            return {
                "employee_code": code,
                "full_name": name,
                "start_date": it.get("start_date"),
                "title_id": title_id,
                "department_id": dept_id,
                "date_of_birth": it.get("date_of_birth"),
                "gender": norm_str(it.get("gender")),
                "national_id": norm_str(it.get("national_id")),
                "id_issue_date": it.get("id_issue_date"),
                "id_issue_place": norm_str(it.get("id_issue_place")),
                "address": norm_str(it.get("address")),
                "phone": norm_str(it.get("phone")),
                "insurance_no": norm_str(it.get("insurance_no")),
                "tax_code": norm_str(it.get("tax_code")),
                "degree": norm_str(it.get("degree")),
                "major": norm_str(it.get("major")),
                "contract1_signed": to_bool(it.get("contract1_signed")),
                "contract1_no": norm_str(it.get("contract1_no")),
                "contract1_sign_date": it.get("contract1_sign_date"),
                "contract1_expire_date": it.get("contract1_expire_date"),
                "contract2_indefinite": to_bool(it.get("contract2_indefinite")),
                "contract2_no": norm_str(it.get("contract2_no")),
                "contract2_sign_date": it.get("contract2_sign_date"),
                "children_count": it.get("children_count"),
                "child_dob_1": it.get("child_dob_1"),
                "child_dob_2": it.get("child_dob_2"),
                "child_dob_3": it.get("child_dob_3"),
                "child_dob_4": it.get("child_dob_4"),
                "note": norm_str(it.get("note")),
            }

        def normalize_db_row(db: dict[str, Any]) -> dict[str, Any]:
            def to_iso(v: Any) -> Any:
                if v is None:
                    return None
                try:
                    return v.isoformat()
                except Exception:
                    return v

            return {
                "employee_code": str(db.get("employee_code") or "").strip(),
                "full_name": str(db.get("full_name") or "").strip(),
                "start_date": to_iso(db.get("start_date")),
                "title_id": db.get("title_id"),
                "department_id": db.get("department_id"),
                "date_of_birth": to_iso(db.get("date_of_birth")),
                "gender": (str(db.get("gender") or "").strip() or None),
                "national_id": (str(db.get("national_id") or "").strip() or None),
                "id_issue_date": to_iso(db.get("id_issue_date")),
                "id_issue_place": (str(db.get("id_issue_place") or "").strip() or None),
                "address": (str(db.get("address") or "").strip() or None),
                "phone": (str(db.get("phone") or "").strip() or None),
                "insurance_no": (str(db.get("insurance_no") or "").strip() or None),
                "tax_code": (str(db.get("tax_code") or "").strip() or None),
                "degree": (str(db.get("degree") or "").strip() or None),
                "major": (str(db.get("major") or "").strip() or None),
                "contract1_signed": bool(int(db.get("contract1_signed") or 0)),
                "contract1_no": (str(db.get("contract1_no") or "").strip() or None),
                "contract1_sign_date": to_iso(db.get("contract1_sign_date")),
                "contract1_expire_date": to_iso(db.get("contract1_expire_date")),
                "contract2_indefinite": bool(int(db.get("contract2_indefinite") or 0)),
                "contract2_no": (str(db.get("contract2_no") or "").strip() or None),
                "contract2_sign_date": to_iso(db.get("contract2_sign_date")),
                "children_count": db.get("children_count"),
                "child_dob_1": to_iso(db.get("child_dob_1")),
                "child_dob_2": to_iso(db.get("child_dob_2")),
                "child_dob_3": to_iso(db.get("child_dob_3")),
                "child_dob_4": to_iso(db.get("child_dob_4")),
                "note": (str(db.get("note") or "").strip() or None),
            }

        inserted = 0
        updated = 0
        skipped = 0
        invalid = 0
        failed = 0

        total = len(rows)
        for idx, it in enumerate(rows, start=1):
            payload = norm_payload(it)
            code = str((payload or {}).get("employee_code") or "").strip()

            if not payload:
                invalid += 1
                if progress_cb:
                    progress_cb(idx, False, code, "Dòng không hợp lệ")
                continue

            try:
                existing = self._repo.get_employee_by_code(code)
                if not existing:
                    self._repo.create_employee(payload)
                    inserted += 1
                    if progress_cb:
                        progress_cb(idx, True, code, "Đã thêm")
                    continue

                if only_new:
                    existing_norm = normalize_db_row(existing)
                    payload_cmp = dict(payload)
                    payload_cmp["contract1_signed"] = bool(payload_cmp.get("contract1_signed"))
                    payload_cmp["contract2_indefinite"] = bool(payload_cmp.get("contract2_indefinite"))

                    changed = False
                    for k, v in payload_cmp.items():
                        if k == "employee_code":
                            continue
                        if existing_norm.get(k) != v:
                            changed = True
                            break

                    if not changed:
                        skipped += 1
                        if progress_cb:
                            progress_cb(idx, True, code, "Bỏ qua (không đổi)")
                        continue

                # only_new=False => always update existing
                self._repo.update_employee(int(existing.get("id")), payload)
                updated += 1
                if progress_cb:
                    progress_cb(idx, True, code, "Đã cập nhật")
            except Exception as exc:
                failed += 1
                if progress_cb:
                    progress_cb(idx, False, code, str(exc))
                continue

        ok_all = failed == 0
        return (
            ok_all,
            f"Tổng: {total} | Thêm mới: {inserted} | Cập nhật: {updated} | Bỏ qua: {skipped} | Không hợp lệ: {invalid} | Thất bại: {failed}",
        )

    def import_csv(self, file_path: str) -> tuple[bool, str]:
        path = Path(file_path)
        if not path.exists() or path.suffix.lower() != ".csv":
            return False, "Vui lòng chọn file .csv hợp lệ."

        def parse_bool(v: Any) -> bool:
            s = str(v or "").strip().lower()
            return s in {"1", "true", "yes", "y", "x"}

        def parse_int(v: Any) -> int | None:
            s = str(v or "").strip()
            if not s:
                return None
            try:
                return int(s)
            except Exception:
                return None

        def parse_date(v: Any) -> str | None:
            s = str(v or "").strip()
            if not s:
                return None
            # already ISO
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    return dt.date().isoformat()
                except Exception:
                    continue
            return None

        items: list[dict[str, Any]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                items.append(
                    {
                        "employee_code": (r.get("employee_code") or "").strip(),
                        "full_name": (r.get("full_name") or "").strip(),
                        "start_date": parse_date(r.get("start_date")),
                        "title_id": parse_int(r.get("title_id")),
                        "department_id": parse_int(r.get("department_id")),
                        "date_of_birth": parse_date(r.get("date_of_birth")),
                        "gender": (r.get("gender") or "").strip() or None,
                        "national_id": (r.get("national_id") or "").strip() or None,
                        "id_issue_date": parse_date(r.get("id_issue_date")),
                        "id_issue_place": (r.get("id_issue_place") or "").strip()
                        or None,
                        "address": (r.get("address") or "").strip() or None,
                        "phone": (r.get("phone") or "").strip() or None,
                        "insurance_no": (r.get("insurance_no") or "").strip() or None,
                        "tax_code": (r.get("tax_code") or "").strip() or None,
                        "degree": (r.get("degree") or "").strip() or None,
                        "major": (r.get("major") or "").strip() or None,
                        "contract1_signed": parse_bool(r.get("contract1_signed")),
                        "contract1_no": (r.get("contract1_no") or "").strip() or None,
                        "contract1_sign_date": parse_date(r.get("contract1_sign_date")),
                        "contract1_expire_date": parse_date(
                            r.get("contract1_expire_date")
                        ),
                        "contract2_indefinite": parse_bool(
                            r.get("contract2_indefinite")
                        ),
                        "contract2_no": (r.get("contract2_no") or "").strip() or None,
                        "contract2_sign_date": parse_date(r.get("contract2_sign_date")),
                        "children_count": parse_int(r.get("children_count")),
                        "child_dob_1": parse_date(r.get("child_dob_1")),
                        "child_dob_2": parse_date(r.get("child_dob_2")),
                        "child_dob_3": parse_date(r.get("child_dob_3")),
                        "child_dob_4": parse_date(r.get("child_dob_4")),
                        "note": (r.get("note") or "").strip() or None,
                    }
                )

        affected, skipped = self._repo.upsert_many(items)
        return (
            True,
            f"Đã nhập: {affected} dòng. Bỏ qua: {skipped} dòng (thiếu Mã NV/Họ tên).",
        )

    def create_employee(self, data: dict[str, Any]) -> tuple[bool, str, int | None]:
        code = str(data.get("employee_code") or "").strip()
        name = str(data.get("full_name") or "").strip()

        if not code:
            return False, "Vui lòng nhập Mã NV.", None
        if not code.isdigit():
            return False, "Mã NV chỉ gồm số.", None
        if len(code) > 5:
            return False, "Mã NV tối đa 5 chữ số.", None
        code = code.zfill(5)

        if not name:
            return False, "Vui lòng nhập Họ và tên.", None

        payload = dict(data)
        payload["employee_code"] = code
        payload["full_name"] = name

        try:
            new_id = self._repo.create_employee(payload)
            return True, "Đã thêm nhân viên.", new_id
        except Exception as exc:
            if "1062" in str(exc) or "Duplicate" in str(exc):
                return False, "Mã NV đã tồn tại.", None
            raise

    def update_employee(
        self, employee_id: int, data: dict[str, Any]
    ) -> tuple[bool, str]:
        code = str(data.get("employee_code") or "").strip()
        name = str(data.get("full_name") or "").strip()

        if not code:
            return False, "Vui lòng nhập Mã NV."
        if not code.isdigit():
            return False, "Mã NV chỉ gồm số."
        if len(code) > 5:
            return False, "Mã NV tối đa 5 chữ số."
        code = code.zfill(5)

        if not name:
            return False, "Vui lòng nhập Họ và tên."

        payload = dict(data)
        payload["employee_code"] = code
        payload["full_name"] = name

        try:
            affected = self._repo.update_employee(int(employee_id), payload)
            if affected <= 0:
                return False, "Không tìm thấy nhân viên để cập nhật."
            return True, "Đã cập nhật thông tin."
        except Exception as exc:
            if "1062" in str(exc) or "Duplicate" in str(exc):
                return False, "Mã NV đã tồn tại."
            raise

    def delete_employee(self, employee_id: int) -> tuple[bool, str]:
        affected = self._repo.delete_employee(int(employee_id))
        if affected <= 0:
            return False, "Không tìm thấy nhân viên để xóa."
        return True, "Đã xóa nhân viên."
