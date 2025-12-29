"""services.import_shift_attendance_services

Service cho tính năng "Import dữ liệu chấm công" -> ghi vào attendance_audit.

Luồng:
- export_shift_attendance_template_xlsx: tạo file mẫu
- read_shift_attendance_from_xlsx: đọc Excel -> list rows để preview
- import_shift_attendance_rows: áp dụng vào DB với rules overwrite/skip

Rules overwrite/skip (theo yêu cầu):
- Nếu dòng audit hiện có có import_locked = 0 (nguồn download/sync): overwrite luôn.
- Nếu import_locked = 1 (đã từng import): so sánh tất cả field -> chỉ update khi có thay đổi, không đổi thì SKIP.

Ghi chú:
- File Excel mẫu/preview theo đúng cột MainContent2 (không có attendance_code/device_no).
- Khi import: nếu đã có dữ liệu audit theo (employee_code, work_date) thì dùng (attendance_code, device_no) hiện có để upsert.
- Nếu chưa có: sẽ cố gắng map attendance_code = employees.mcc_code (nếu có) else employee_code; device_no mặc định = 1.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable

from repository.import_shift_attendance_repository import (
    ImportShiftAttendanceRepository,
)
from repository.schedule_work_repository import ScheduleWorkRepository
from repository.shift_attendance_maincontent2_repository import (
    ShiftAttendanceMainContent2Repository,
)
from services.shift_attendance_maincontent2_services import (
    ShiftAttendanceMainContent2Service,
)


logger = logging.getLogger(__name__)


@dataclass
class ImportShiftAttendanceResult:
    ok: bool
    message: str
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0


class ImportShiftAttendanceService:
    def __init__(
        self, repository: ImportShiftAttendanceRepository | None = None
    ) -> None:
        self._repo = repository or ImportShiftAttendanceRepository()

    @staticmethod
    def _weekday_label(d: date) -> str:
        # 0=Mon..6=Sun
        w = int(d.weekday())
        return (
            "Thứ 2"
            if w == 0
            else (
                "Thứ 3"
                if w == 1
                else (
                    "Thứ 4"
                    if w == 2
                    else (
                        "Thứ 5"
                        if w == 3
                        else "Thứ 6" if w == 4 else "Thứ 7" if w == 5 else "Chủ nhật"
                    )
                )
            )
        )

    @staticmethod
    def _normalize_employee_code(
        value: Any,
        *,
        pad_width: int | None = None,
    ) -> str:
        """Normalize employee_code read from Excel.

        Excel often drops leading zeros and returns numeric types. If the file
        contains other codes with leading zeros, we pad numeric-only codes to
        the same width so matching/upsert works (e.g. 4 -> 00004).
        """

        if value is None:
            return ""

        # Handle numeric cells (int/float) that represent codes.
        if isinstance(value, bool):
            s = ""
        elif isinstance(value, int):
            s = str(value)
        elif isinstance(value, float):
            try:
                if float(value).is_integer():
                    s = str(int(value))
                else:
                    s = str(value)
            except Exception:
                s = str(value)
        else:
            s = str(value)

        s = s.strip()
        if not s:
            return ""

        if pad_width and pad_width > 0 and s.isdigit() and len(s) < int(pad_width):
            return s.zfill(int(pad_width))

        return s

    @staticmethod
    def export_shift_attendance_template_xlsx(file_path: str) -> tuple[bool, str]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng chọn đường dẫn lưu file mẫu."
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện ghi Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
            )

        headers = [
            "Mã nv",
            "Tên nhân viên",
            "Ngày",
            "Thứ",
            "Vào 1",
            "Ra 1",
            "Vào 2",
            "Ra 2",
            "Vào 3",
            "Ra 3",
            "Trễ",
            "Sớm",
            "Giờ",
            "Công",
            "KH",
            "Giờ +",
            "Công +",
            "KH +",
            "TC1",
            "TC2",
            "TC3",
            "Lịch NV",
            "Ca",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "ChamCong"

        header_font = Font(bold=True)
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font

        for col_idx, label in enumerate(headers, start=1):
            width = max(12, min(28, len(label) + 6))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                width
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

        # Auto-open exported template (best-effort)
        try:
            import os

            if os.name == "nt":
                os.startfile(str(path))  # type: ignore[attr-defined]
        except Exception:
            pass
        return True, f"Đã tạo file mẫu: {path}"

    def read_shift_attendance_from_xlsx(
        self, file_path: str
    ) -> tuple[bool, str, list[dict[str, Any]]]:
        path = Path(file_path)
        if not str(path).strip():
            return False, "Vui lòng nhập đường dẫn file Excel.", []
        if not path.exists() or path.suffix.lower() != ".xlsx":
            return False, "Vui lòng chọn file .xlsx hợp lệ.", []

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return (
                False,
                "Thiếu thư viện đọc Excel. Vui lòng cài 'openpyxl' trong môi trường Python.",
                [],
            )

        def norm_header(s: Any) -> str:
            raw = str(s or "").strip()
            raw = re.sub(r"\s+", " ", raw)
            return raw

        def norm_header_key(s: Any) -> str:
            raw = str(s or "").strip().lower()
            raw = re.sub(r"\s+", " ", raw)
            raw = raw.replace("đ", "d").replace("Đ", "d")
            raw = raw.replace("\ufffd", "")
            raw = unicodedata.normalize("NFKD", raw)
            raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
            raw = re.sub(r"[^0-9a-z ]+", "", raw)
            raw = raw.replace(" ", "")
            return raw

        header_to_key: dict[str, str] = {
            # keys
            "employee_code": "employee_code",
            "full_name": "full_name",
            "date": "work_date",
            "work_date": "work_date",
            "weekday": "weekday",
            "schedule": "schedule",
            "shift_code": "shift_code",
            "in_1": "in_1",
            "out_1": "out_1",
            "in_2": "in_2",
            "out_2": "out_2",
            "in_3": "in_3",
            "out_3": "out_3",
            "late": "late",
            "early": "early",
            "hours": "hours",
            "work": "work",
            "leave": "leave",
            "hours_plus": "hours_plus",
            "work_plus": "work_plus",
            "leave_plus": "leave_plus",
            "tc1": "tc1",
            "tc2": "tc2",
            "tc3": "tc3",
            "in_1_symbol": "in_1_symbol",
            "symbol": "in_1_symbol",
            # Vietnamese (MainContent2)
            "Mã nv": "employee_code",
            "Mã NV": "employee_code",
            "Tên nhân viên": "full_name",
            "Họ và tên": "full_name",
            "Ngày": "work_date",
            "Ngay": "work_date",
            "Ngày công": "work_date",
            "Ngay cong": "work_date",
            "Thứ": "weekday",
            "Ký hiệu": "in_1_symbol",
            "Ky hieu": "in_1_symbol",
            "Kí hiệu": "in_1_symbol",
            "Ca": "shift_code",
            "Lịch NV": "schedule",
            "Lịch làm việc": "schedule",
            "Vào 1": "in_1",
            "Ra 1": "out_1",
            "Vào 2": "in_2",
            "Ra 2": "out_2",
            "Vào 3": "in_3",
            "Ra 3": "out_3",
            "Trễ": "late",
            "Sớm": "early",
            "Giờ": "hours",
            "Công": "work",
            "KH": "leave",
            "Giờ +": "hours_plus",
            "Công +": "work_plus",
            "KH +": "leave_plus",
            "TC1": "tc1",
            "TC2": "tc2",
            "TC3": "tc3",
            # Backward-compat: older template used 'Tổng' column
            "Tổng": "schedule",
            # ignored
            "STT": "stt",
            "ID": "id",
            "__check": "__check",
            "": "",
        }

        def parse_date(v: Any) -> str | None:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v.date().isoformat()
            if isinstance(v, date):
                return v.isoformat()
            s = str(v or "").strip()
            if not s:
                return None
            # dd/MM/yyyy
            try:
                if "/" in s and len(s.split("/")) == 3:
                    dd, mm, yy = s.split("/")
                    return date(int(yy), int(mm), int(dd)).isoformat()
            except Exception:
                pass
            try:
                return date.fromisoformat(s).isoformat()
            except Exception:
                return None

        def parse_time(v: Any) -> time | None:
            if v is None:
                return None
            if isinstance(v, time):
                return v
            if isinstance(v, datetime):
                return v.time().replace(microsecond=0)
            s = str(v or "").strip()
            if not s:
                return None
            s = s.replace(".", ":")
            parts = s.split(":")
            try:
                if len(parts) == 2:
                    hh, mm = parts
                    return time(int(hh), int(mm), 0)
                if len(parts) >= 3:
                    hh, mm, ss = parts[:3]
                    return time(int(hh), int(mm), int(float(ss)))
            except Exception:
                return None
            return None

        def extract_symbol_token(v: Any) -> str:
            """Extract a short symbol code from a cell value.

            Some files place attendance symbols in time columns (e.g. 'V', 'KR').
            We treat these as in_1_symbol so the import doesn't drop them.
            """

            if v is None:
                return ""
            if isinstance(v, (time, datetime)):
                return ""
            s = str(v or "").strip()
            if not s:
                return ""
            s2 = s.replace(" ", "")
            # If it looks like a time, it's not a symbol.
            if ":" in s2 or "." in s2:
                return ""
            # Reject obvious numbers.
            if s2.isdigit():
                return ""
            # Keep short alpha-numeric codes like V, KR, OFF, KV...
            if len(s2) > 12:
                return ""
            # Must contain at least one letter.
            if not any(ch.isalpha() for ch in s2):
                return ""
            return s2

        def parse_decimal(v: Any) -> Decimal | None:
            if v is None:
                return None
            if isinstance(v, Decimal):
                return v
            if isinstance(v, (int, float)):
                try:
                    return Decimal(str(v))
                except Exception:
                    return None
            s = str(v or "").strip()
            if not s:
                return None
            s = s.replace(" ", "")
            # Vietnamese style: 1,5
            if s.count(",") == 1 and s.count(".") == 0:
                s = s.replace(",", ".")
            try:
                return Decimal(s)
            except Exception:
                return None

        def parse_decimal_or_symbol(v: Any) -> Decimal | str | None:
            """Parse a cell that may contain either a number or a symbol.

            Some templates/users put symbols like 'V' into KH / KH+ columns.
            The Shift Attendance UI logic also treats leave/leave_plus as a symbol
            in some cases.
            """

            if v is None:
                return None
            # First try numeric
            dv = parse_decimal(v)
            if dv is not None:
                return dv
            # Then try symbol token
            sym = extract_symbol_token(v)
            return sym or None

        wb = load_workbook(str(path), data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return False, "File Excel trống.", []

        headers = [norm_header(h) for h in list(header_row or [])]
        header_to_key_lower = {
            str(k).strip().lower(): v for k, v in header_to_key.items()
        }
        header_to_key_norm = {norm_header_key(k): v for k, v in header_to_key.items()}

        col_keys: list[str | None] = []
        for h in headers:
            key = header_to_key.get(h)
            if key is None:
                key = header_to_key_lower.get(str(h or "").strip().lower())
            if key is None:
                key = header_to_key_norm.get(norm_header_key(h))
            col_keys.append(key)

        unknown_headers = [
            headers[i]
            for i, k in enumerate(col_keys)
            if k is None and str(headers[i] or "").strip()
        ]
        if unknown_headers:
            # Không fail; chỉ cảnh báo nhẹ trong message
            pass

        out: list[dict[str, Any]] = []
        for r in rows_iter:
            if r is None:
                continue
            item: dict[str, Any] = {}
            empty = True
            for idx, raw in enumerate(list(r)):
                key = col_keys[idx] if idx < len(col_keys) else None
                if not key or key in {"id", "stt", "__check"}:
                    continue

                if raw is not None and str(raw).strip() != "":
                    empty = False

                if key == "work_date":
                    item[key] = parse_date(raw)
                elif key in {"in_1", "out_1", "in_2", "out_2", "in_3", "out_3"}:
                    tv = parse_time(raw)
                    if tv is not None:
                        item[key] = tv
                    else:
                        # If user/device put a symbol code in a time column, keep it.
                        sym = extract_symbol_token(raw)
                        if sym:
                            cur_sym = str(item.get("in_1_symbol") or "").strip()
                            if not cur_sym:
                                item["in_1_symbol"] = sym
                            else:
                                # Avoid duplicates; keep stable order.
                                parts = [p for p in cur_sym.split("|") if p]
                                if sym not in parts:
                                    item["in_1_symbol"] = "|".join(parts + [sym])
                        item[key] = None
                elif key in {
                    "hours",
                    "work",
                    "hours_plus",
                    "work_plus",
                    # leave/leave_plus handled separately (can be symbol)
                }:
                    item[key] = parse_decimal(raw)
                elif key in {"leave", "leave_plus"}:
                    item[key] = parse_decimal_or_symbol(raw)
                else:
                    s = str(raw or "").strip()
                    item[key] = s if s else None

            if empty:
                continue

            emp_code = str(item.get("employee_code") or "").strip()
            if emp_code:
                item["employee_code"] = emp_code

            wd = str(item.get("work_date") or "").strip()
            if wd:
                item["work_date"] = wd
                try:
                    d = date.fromisoformat(wd)
                    item.setdefault("weekday", self._weekday_label(d))
                except Exception:
                    pass

            out.append(item)

        msg = f"Đọc thành công {len(out)} dòng."
        if unknown_headers:
            msg += f" (Bỏ qua cột lạ: {', '.join(unknown_headers[:6])}{'...' if len(unknown_headers) > 6 else ''})"

        # Normalize employee_code: if the sheet contains zero-padded numeric codes
        # (e.g. '00010'), pad numeric codes like 4 -> '00004' to the same width.
        try:
            pad_width = 0
            for it in out:
                v = it.get("employee_code")
                s = str(v or "").strip()
                if s.isdigit() and s.startswith("0"):
                    pad_width = max(pad_width, len(s))
            if pad_width > 0:
                for it in out:
                    it["employee_code"] = self._normalize_employee_code(
                        it.get("employee_code"),
                        pad_width=pad_width,
                    )
        except Exception:
            pass

        return True, msg, out

    def import_shift_attendance_rows(
        self,
        rows: list[dict[str, Any]],
        progress_cb: Callable[[int, bool, str, str], None] | None = None,
        report: list[dict[str, Any]] | None = None,
    ) -> ImportShiftAttendanceResult:
        if not rows:
            return ImportShiftAttendanceResult(False, "Không có dữ liệu để cập nhập.")

        # Normalize employee_code first (rows may come from UI edits).
        try:
            pad_width = 0
            for r0 in rows or []:
                s0 = str(r0.get("employee_code") or "").strip()
                if s0.isdigit() and s0.startswith("0"):
                    pad_width = max(pad_width, len(s0))
            if pad_width > 0:
                for r0 in rows or []:
                    r0["employee_code"] = self._normalize_employee_code(
                        r0.get("employee_code"),
                        pad_width=pad_width,
                    )
        except Exception:
            pass

        # Build keys + employee lookup first (needed to map mcc_code -> attendance_code).
        pairs_emp: list[tuple[str, str]] = []
        emp_codes: list[str] = []
        work_dates: list[str] = []
        for r in rows:
            emp_code = str(r.get("employee_code") or "").strip()
            wd = str(r.get("work_date") or "").strip()
            if emp_code and wd:
                pairs_emp.append((emp_code, wd))
                emp_codes.append(emp_code)
            work_dates.append(wd)

        emp_lookup: dict[str, dict[str, Any]] = {}
        try:
            emp_lookup = self._repo.get_employees_by_codes(emp_codes)
        except Exception:
            emp_lookup = {}

        # Existing lookup:
        # - primary: by (employee_code, work_date)
        # - fallback: by (attendance_code=mcc_code, work_date) for legacy downloaded rows
        existing_map_by_emp: dict[tuple[str, str], dict[str, Any]] = {}
        try:
            existing_map_by_emp = self._repo.get_existing_by_employee_code_date(
                pairs_emp
            )
        except Exception:
            existing_map_by_emp = {}

        pairs_att: list[tuple[str, str]] = []
        try:
            for emp_code, wd in pairs_emp:
                emp = emp_lookup.get(str(emp_code).lower())
                mcc = str((emp or {}).get("mcc_code") or "").strip()
                if mcc:
                    pairs_att.append((mcc, wd))
        except Exception:
            pairs_att = []

        existing_map_by_att: dict[tuple[str, str], dict[str, Any]] = {}
        try:
            if pairs_att:
                existing_map_by_att = self._repo.get_existing_by_attendance_code_date(
                    pairs_att
                )
        except Exception:
            existing_map_by_att = {}

        def _get_existing_row(emp_code: str, wd: str) -> dict[str, Any] | None:
            k = (str(emp_code or "").strip(), str(wd or "").strip())
            if k[0] and k[1]:
                r0 = existing_map_by_emp.get(k)
                if r0:
                    return r0
            emp = emp_lookup.get(str(emp_code or "").lower())
            mcc = str((emp or {}).get("mcc_code") or "").strip()
            if mcc and k[1]:
                return existing_map_by_att.get((mcc, k[1]))
            return None

        def _iter_dates_inclusive(from_iso: str, to_iso: str) -> list[str]:
            try:
                d0 = date.fromisoformat(str(from_iso))
                d1 = date.fromisoformat(str(to_iso))
            except Exception:
                return []
            if d1 < d0:
                d0, d1 = d1, d0
            days = (d1 - d0).days
            # Guardrail: importing an extremely wide range can create too many rows.
            if days > 180:
                return []
            out: list[str] = []
            cur = d0
            while cur <= d1:
                out.append(cur.isoformat())
                cur = cur.fromordinal(cur.toordinal() + 1)
            return out

        def _get_existing_pairs(pairs0: list[tuple[str, str]]) -> set[tuple[str, str]]:
            existing: set[tuple[str, str]] = set()
            if not pairs0:
                return existing

            # Reverse map: mcc_code(attendance_code) -> employee_code for the imported population.
            mcc_to_emp: dict[str, str] = {}
            try:
                for ec0 in emp_codes:
                    emp0 = emp_lookup.get(str(ec0 or "").lower())
                    mcc0 = str((emp0 or {}).get("mcc_code") or "").strip()
                    if mcc0:
                        mcc_to_emp[mcc0] = str(ec0)
            except Exception:
                mcc_to_emp = {}

            # Chunk to avoid overly long IN lists.
            step = 800
            for i0 in range(0, len(pairs0), step):
                part = pairs0[i0 : i0 + step]
                # 1) employee_code based
                try:
                    m1 = self._repo.get_existing_by_employee_code_date(part)
                except Exception:
                    m1 = {}
                for k in (m1 or {}).keys():
                    try:
                        ec, wd = k
                        if ec and wd:
                            existing.add((str(ec).strip(), str(wd).strip()))
                    except Exception:
                        continue

                # 2) attendance_code based (mcc_code) to catch legacy downloaded rows
                try:
                    part2: list[tuple[str, str]] = []
                    for ec, wd in part:
                        emp = emp_lookup.get(str(ec or "").lower())
                        mcc = str((emp or {}).get("mcc_code") or "").strip()
                        if mcc and wd:
                            part2.append((mcc, wd))
                    if part2:
                        m2 = self._repo.get_existing_by_attendance_code_date(part2)
                    else:
                        m2 = {}
                except Exception:
                    m2 = {}
                if m2:
                    # Map back to (employee_code, work_date) so callers can skip placeholders
                    for (att_code, wd), row in (m2 or {}).items():
                        try:
                            ec2 = str((row or {}).get("employee_code") or "").strip()
                            if not ec2:
                                ec2 = str(
                                    mcc_to_emp.get(str(att_code or "").strip()) or ""
                                ).strip()
                            if ec2 and wd:
                                existing.add((ec2, str(wd).strip()))
                        except Exception:
                            continue
            return existing

        def _build_missing_day_payloads(
            *,
            from_date_iso: str,
            to_date_iso: str,
        ) -> tuple[list[dict[str, Any]], set[tuple[str, str]]]:
            """Create placeholder rows for days without any attendance_audit row.

            Returns (payloads, affected_pairs).
            """

            date_list = _iter_dates_inclusive(from_date_iso, to_date_iso)
            if not date_list:
                return ([], set())

            # Only generate placeholders for employees that can be mapped to employee_id.
            emp_id_by_code: dict[str, int] = {}
            for code in emp_codes:
                c = str(code or "").strip()
                if not c:
                    continue
                emp = emp_lookup.get(c.lower())
                if not emp or emp.get("id") is None:
                    continue
                try:
                    eid = int(emp.get("id"))
                except Exception:
                    continue
                if eid > 0:
                    emp_id_by_code[c] = eid

            if not emp_id_by_code:
                return ([], set())

            # Strongest existence check: any row for (employee_id, work_date) should block placeholder.
            existing_empid_pairs: set[tuple[int, str]] = set()
            try:
                all_pairs_empid: list[tuple[int, str]] = []
                for _ec, _eid in emp_id_by_code.items():
                    for d_iso in date_list:
                        all_pairs_empid.append((int(_eid), str(d_iso)))
                # Chunk to avoid huge IN lists.
                step2 = 800
                for j0 in range(0, len(all_pairs_empid), step2):
                    part_ids = all_pairs_empid[j0 : j0 + step2]
                    try:
                        existing_empid_pairs |= (
                            self._repo.get_existing_employee_id_date_pairs(part_ids)
                        )
                    except Exception:
                        pass
            except Exception:
                existing_empid_pairs = set()

            # Preload schedule names for each date for all employee_ids (batch per date).
            schedule_by_emp_date2: dict[tuple[int, str], str] = {}
            try:
                sched_repo2 = ScheduleWorkRepository()
                ids_all = list(dict.fromkeys(list(emp_id_by_code.values())))
                for d_iso in date_list:
                    try:
                        m2 = sched_repo2.get_employee_schedule_name_map(
                            employee_ids=ids_all,
                            on_date=str(d_iso),
                        )
                    except Exception:
                        m2 = {}
                    for eid2, name2 in (m2 or {}).items():
                        sname2 = str(name2 or "").strip()
                        if sname2:
                            schedule_by_emp_date2[(int(eid2), str(d_iso))] = sname2
            except Exception:
                schedule_by_emp_date2 = {}

            # Determine which (employee_code, date) pairs already exist in DB.
            all_pairs: list[tuple[str, str]] = []
            for ec, _eid in emp_id_by_code.items():
                for d_iso in date_list:
                    all_pairs.append((str(ec), str(d_iso)))
            existing_pairs = _get_existing_pairs(all_pairs)

            payloads: list[dict[str, Any]] = []
            affected: set[tuple[str, str]] = set()
            for ec, eid in emp_id_by_code.items():
                emp = emp_lookup.get(ec.lower())
                name = str(
                    (emp or {}).get("full_name") or (emp or {}).get("name_on_mcc") or ""
                ).strip()
                mcc = str((emp or {}).get("mcc_code") or "").strip()

                for d_iso in date_list:
                    k = (str(ec).strip(), str(d_iso).strip())
                    # If ANY row already exists for this employee_id+date (any device/attendance_code), skip placeholder.
                    try:
                        if (int(eid), str(d_iso).strip()) in existing_empid_pairs:
                            continue
                    except Exception:
                        pass
                    if k in existing_pairs:
                        continue

                    schedule_name = str(
                        schedule_by_emp_date2.get((int(eid), str(d_iso))) or ""
                    ).strip()
                    weekday = ""
                    try:
                        weekday = self._weekday_label(date.fromisoformat(str(d_iso)))
                    except Exception:
                        weekday = ""

                    payloads.append(
                        {
                            "attendance_code": mcc or str(ec),
                            "device_no": 1,
                            "device_id": None,
                            "device_name": "",
                            "employee_id": int(eid),
                            "employee_code": str(ec),
                            "full_name": name or None,
                            "work_date": str(d_iso),
                            "weekday": weekday or None,
                            "schedule": schedule_name or None,
                            "shift_code": None,
                            "in_1_symbol": None,
                            "in_1": None,
                            "out_1": None,
                            "in_2": None,
                            "out_2": None,
                            "in_3": None,
                            "out_3": None,
                            "late": None,
                            "early": None,
                            "hours": None,
                            "work": None,
                            "leave": None,
                            "hours_plus": None,
                            "work_plus": None,
                            "leave_plus": None,
                            "tc1": None,
                            "tc2": None,
                            "tc3": None,
                            # IMPORTANT: allow future downloads to overwrite placeholders.
                            "import_locked": 0,
                        }
                    )
                    affected.add(k)

            return (payloads, affected)

        # Schedule snapshot for each (employee_id, work_date).
        # If Excel doesn't provide schedule, we resolve current assignment on that date
        # and store it into attendance_audit to "chốt công".
        schedule_by_emp_date: dict[tuple[int, str], str] = {}
        try:
            by_date_emp_ids: dict[str, list[int]] = {}
            for r0 in rows:
                wd0 = str(r0.get("work_date") or "").strip()
                if not wd0:
                    continue
                ec0 = str(r0.get("employee_code") or "").strip()
                if not ec0:
                    continue
                emp0 = emp_lookup.get(ec0.lower())
                if not emp0 or emp0.get("id") is None:
                    continue
                try:
                    eid0 = int(emp0.get("id"))
                except Exception:
                    continue
                if eid0 <= 0:
                    continue
                by_date_emp_ids.setdefault(wd0, []).append(eid0)

            sched_repo = ScheduleWorkRepository()
            for wd0, emp_ids0 in by_date_emp_ids.items():
                ids0 = list(dict.fromkeys([int(x) for x in emp_ids0 if int(x) > 0]))
                if not ids0:
                    continue
                m0 = sched_repo.get_employee_schedule_name_map(
                    employee_ids=ids0,
                    on_date=str(wd0),
                )
                for eid0, name0 in (m0 or {}).items():
                    sname0 = str(name0 or "").strip()
                    if sname0:
                        schedule_by_emp_date[(int(eid0), str(wd0))] = sname0
        except Exception:
            schedule_by_emp_date = {}

        def to_time_str(t: Any) -> str | None:
            if t is None:
                return None
            if isinstance(t, time):
                return t.replace(microsecond=0).strftime("%H:%M:%S")
            return str(t)

        def to_dec_str(v: Any) -> str | None:
            if v is None:
                return None
            if isinstance(v, Decimal):
                return format(v, "f")
            try:
                return format(Decimal(str(v)), "f")
            except Exception:
                s = str(v or "").strip()
                return s if s else None

        compare_keys = [
            "employee_code",
            "full_name",
            "work_date",
            "weekday",
            "schedule",
            "in_1_symbol",
            "in_1",
            "out_1",
            "in_2",
            "out_2",
            "in_3",
            "out_3",
            "late",
            "early",
            "hours",
            "work",
            "leave",
            "hours_plus",
            "work_plus",
            "leave_plus",
            "tc1",
            "tc2",
            "tc3",
            "shift_code",
        ]

        def normalize_for_compare(k: str, v: Any) -> Any:
            if k in {"in_1", "out_1", "in_2", "out_2", "in_3", "out_3"}:
                return to_time_str(v)
            if k in {
                "hours",
                "work",
                "leave",
                "hours_plus",
                "work_plus",
                "leave_plus",
            }:
                return to_dec_str(v)
            if k == "work_date":
                if v is None:
                    return None
                if isinstance(v, (date, datetime)):
                    return (v.date() if isinstance(v, datetime) else v).isoformat()
                return str(v)
            s = str(v or "").strip()
            return s if s else None

        def add_report(
            *,
            idx: int,
            code: str,
            result: str,
            action: str,
            message: str,
        ) -> None:
            if report is None:
                return
            report.append(
                {
                    "index": int(idx),
                    "employee_code": str(code or "").strip(),
                    "result": str(result or ""),
                    "action": str(action or ""),
                    "message": str(message or ""),
                }
            )

        upsert_payloads: list[dict[str, Any]] = []

        inserted = 0
        updated = 0
        skipped = 0
        failed = 0
        inserted_missing = 0

        total = len(rows)
        for i, raw in enumerate(rows, start=1):
            emp_code = str(raw.get("employee_code") or "").strip()
            wd = str(raw.get("work_date") or "").strip()

            if not emp_code or not wd:
                failed += 1
                add_report(
                    idx=i,
                    code=emp_code or "(không mã)",
                    result="INVALID",
                    action="SKIP_INVALID",
                    message="Thiếu Mã nv hoặc Ngày.",
                )
                if progress_cb:
                    progress_cb(i, False, emp_code or "(không mã)", "Thiếu dữ liệu")
                continue

            existing = _get_existing_row(emp_code, wd)
            import_locked = int(existing.get("import_locked") or 0) if existing else 0

            # If not found in DB, we allow INSERT (create new row) from Excel import.

            # Determine whether changed (only when import_locked=1)
            changed = True
            if existing and import_locked == 1:
                changed = False
                for k in compare_keys:
                    new_v = normalize_for_compare(k, raw.get(k))
                    old_v = normalize_for_compare(k, existing.get(k))
                    if new_v != old_v:
                        changed = True
                        break

            if existing and import_locked == 1 and not changed:
                skipped += 1
                add_report(
                    idx=i,
                    code=emp_code,
                    result="SKIPPED",
                    action="SKIP_NO_CHANGE",
                    message="Không thay đổi.",
                )
                if progress_cb:
                    progress_cb(i, True, emp_code, "Bỏ qua (không đổi)")
                continue

            # Build payload for upsert
            payload: dict[str, Any] = {}

            # Resolve base keys for unique
            if existing:
                payload["attendance_code"] = (
                    str(existing.get("attendance_code") or "").strip() or emp_code
                )
                payload["device_no"] = int(existing.get("device_no") or 1)
                payload["device_id"] = existing.get("device_id")
                payload["device_name"] = existing.get("device_name")
            else:
                emp = emp_lookup.get(emp_code.lower())
                mcc = str((emp or {}).get("mcc_code") or "").strip()
                payload["attendance_code"] = mcc or emp_code
                payload["device_no"] = 1
                payload["device_id"] = None
                payload["device_name"] = ""

            emp = emp_lookup.get(emp_code.lower())
            payload["employee_id"] = (
                int(emp.get("id")) if emp and emp.get("id") is not None else None
            )
            payload["employee_code"] = emp_code

            name = str(raw.get("full_name") or "").strip()
            if not name and emp:
                name = str(emp.get("full_name") or emp.get("name_on_mcc") or "").strip()
            payload["full_name"] = name or None

            payload["work_date"] = wd
            # weekday: use provided, else compute
            weekday = str(raw.get("weekday") or "").strip()
            if not weekday:
                try:
                    weekday = self._weekday_label(date.fromisoformat(wd))
                except Exception:
                    weekday = ""
            payload["weekday"] = weekday or None

            # Times + fields
            for k in ["in_1", "out_1", "in_2", "out_2", "in_3", "out_3"]:
                payload[k] = raw.get(k)

            for k in [
                "late",
                "early",
                "tc1",
                "tc2",
                "tc3",
            ]:
                v = str(raw.get(k) or "").strip()
                payload[k] = v or None

            for k in [
                "hours",
                "work",
                "leave",
                "hours_plus",
                "work_plus",
                "leave_plus",
            ]:
                payload[k] = raw.get(k)

            # schedule is a display field (varchar)
            schedule = str(raw.get("schedule") or "").strip()
            if not schedule:
                try:
                    eid = payload.get("employee_id")
                    if eid is not None:
                        schedule = schedule_by_emp_date.get((int(eid), wd), "")
                except Exception:
                    schedule = ""
            payload["schedule"] = schedule or None

            # Optional symbol column (e.g. 'Ký hiệu')
            sym = str(raw.get("in_1_symbol") or "").strip()
            payload["in_1_symbol"] = sym or None

            # KH/KH+ columns may carry a symbol in some user templates (e.g., 'V').
            # DB columns are numeric on many installs; never push string into them.
            def _merge_symbol_into_payload(val: Any) -> None:
                try:
                    s0 = str(val or "").strip()
                except Exception:
                    s0 = ""
                if not s0:
                    return
                cur = str(payload.get("in_1_symbol") or "").strip()
                if not cur:
                    payload["in_1_symbol"] = s0
                    return
                parts = [p.strip() for p in cur.split("|") if p.strip()]
                if s0 not in parts:
                    payload["in_1_symbol"] = "|".join(parts + [s0])

            for k_sym in ("leave", "leave_plus"):
                v_sym = raw.get(k_sym)
                if isinstance(v_sym, str) and str(v_sym).strip():
                    _merge_symbol_into_payload(v_sym)
                    payload[k_sym] = None

            # shift_code (Ca) is optional; upload only when provided.
            shift_code = str(raw.get("shift_code") or "").strip()
            payload["shift_code"] = shift_code or None

            # Mark as imported
            payload["import_locked"] = 1

            # Decide action label
            if existing:
                action = (
                    "OVERWRITE_DOWNLOAD" if import_locked == 0 else "UPDATE_CHANGED"
                )
                updated += 1
            else:
                action = "INSERT"
                inserted += 1

            upsert_payloads.append(payload)
            add_report(
                idx=i,
                code=emp_code,
                result="SUCCESS",
                action=action,
                message="Sẽ cập nhập" if existing else "Sẽ thêm",
            )
            if progress_cb:
                progress_cb(i, True, emp_code, action)

        # Fill missing days in DB so Shift Attendance can show V/KV/OFF/Lễ correctly.
        missing_payloads: list[dict[str, Any]] = []
        missing_pairs: set[tuple[str, str]] = set()
        try:
            dates2 = [d for d in work_dates if d]
            from_date2 = min(dates2) if dates2 else None
            to_date2 = max(dates2) if dates2 else None
            if from_date2 and to_date2:
                missing_payloads, missing_pairs = _build_missing_day_payloads(
                    from_date_iso=str(from_date2),
                    to_date_iso=str(to_date2),
                )
        except Exception:
            missing_payloads, missing_pairs = ([], set())

        if missing_payloads:
            try:
                self._repo.upsert_import_rows(missing_payloads)
                inserted_missing = int(len(missing_payloads))
            except Exception:
                # Non-fatal: importing actual rows still succeeded.
                logger.exception(
                    "Không thể tạo dòng placeholder cho ngày thiếu dữ liệu"
                )
                inserted_missing = 0

        # Execute upserts in one batch
        try:
            self._repo.upsert_import_rows(upsert_payloads)
        except Exception as exc:
            # Mark remaining as failed (best-effort)
            logger.exception("Import attendance_audit thất bại")
            return ImportShiftAttendanceResult(
                False,
                f"Không thể cập nhập CSDL: {exc}",
                inserted=inserted,
                updated=updated,
                skipped=skipped,
                failed=(failed + max(0, len(upsert_payloads))),
            )

        # After import: recompute derived/computed columns and persist back to DB.
        # IMPORTANT: When user re-uploads, they expect recalculation even if the import
        # decides to SKIP because values look unchanged. Therefore, recompute must be
        # based on the uploaded Excel rows, not only on rows that were upserted.
        try:
            if rows:
                # Match affected rows by BOTH employee_code and attendance_code.
                # Some DBs store/compare keys differently (e.g. '00004' vs '4').
                affected_keys: set[tuple[str, str]] = set()  # (code, work_date)
                provided_shift_code_by_key: dict[tuple[str, str], str] = {}
                att_codes: list[str] = []
                emp_ids: list[int] = []
                dates: list[str] = []

                def _add_key(code0: Any, wd0: Any) -> None:
                    c0 = str(code0 or "").strip()
                    d0 = str(wd0 or "").strip()
                    if c0 and d0:
                        affected_keys.add((c0, d0))

                for p in rows:
                    wd = str(p.get("work_date") or "").strip()
                    if wd:
                        dates.append(wd)

                    ec = str(p.get("employee_code") or "").strip()
                    ac2 = str(p.get("attendance_code") or "").strip()
                    _add_key(ec, wd)
                    _add_key(ac2, wd)

                    sc = str(p.get("shift_code") or "").strip()
                    if sc:
                        if ec and wd:
                            provided_shift_code_by_key[(ec, wd)] = sc
                        if ac2 and wd:
                            provided_shift_code_by_key[(ac2, wd)] = sc

                    if ac2:
                        att_codes.append(ac2)
                    # Also include employee_code in attendance_codes filter so the repo
                    # can match either a.attendance_code OR a.employee_code.
                    if ec:
                        att_codes.append(ec)

                    try:
                        eid = p.get("employee_id")
                        if eid is not None:
                            emp_ids.append(int(eid))
                    except Exception:
                        pass

                att_codes = list(dict.fromkeys([s for s in att_codes if s]))
                emp_ids = list(dict.fromkeys(emp_ids))
                dates = [d for d in dates if d]
                from_date = min(dates) if dates else None
                to_date = max(dates) if dates else None

                # Include missing-day placeholders in recompute persistence.
                try:
                    if missing_pairs:
                        for (ec2, wd2) in missing_pairs:
                            _add_key(ec2, wd2)
                except Exception:
                    pass

                if from_date and to_date and affected_keys:
                    calc_service = ShiftAttendanceMainContent2Service()
                    rows_calc = calc_service.list_attendance_audit_arranged(
                        from_date=from_date,
                        to_date=to_date,
                        employee_ids=emp_ids or None,
                        attendance_codes=att_codes or None,
                        recompute_import_locked=True,
                        overwrite_import_locked_computed=True,
                    )

                    # Persist only exact affected rows from this upload.
                    pending_updates: list[dict[str, Any]] = []

                    def _to_num(v: Any) -> Any:
                        if v is None:
                            return None
                        if isinstance(v, (int, float, Decimal)):
                            return v
                        s = str(v).strip()
                        if not s:
                            return None
                        try:
                            return Decimal(s)
                        except Exception:
                            try:
                                return float(s)
                            except Exception:
                                return None

                    for r in rows_calc or []:
                        try:
                            ec = str(r.get("employee_code") or "").strip()
                            ac = str(r.get("attendance_code") or "").strip()
                            wd = str(r.get("date") or r.get("work_date") or "").strip()
                            if not wd:
                                continue

                            # Accept match by either code.
                            if (ec and (ec, wd) in affected_keys) or (
                                ac and (ac, wd) in affected_keys
                            ):
                                pass
                            else:
                                continue

                            work_v = _to_num(r.get("work"))
                            work_plus_v = _to_num(r.get("work_plus"))
                            total_v = None
                            if (work_v is not None) or (work_plus_v is not None):
                                try:
                                    total_v = (work_v or 0) + (work_plus_v or 0)
                                except Exception:
                                    total_v = None

                            pending_updates.append(
                                {
                                    "id": r.get("id"),
                                    "work_date": wd,
                                    "late": r.get("late"),
                                    "early": r.get("early"),
                                    "hours": r.get("hours"),
                                    "work": work_v,
                                    "hours_plus": r.get("hours_plus"),
                                    "work_plus": work_plus_v,
                                    "tc1": r.get("tc1"),
                                    "tc2": r.get("tc2"),
                                    "tc3": r.get("tc3"),
                                    "total": total_v,
                                    "schedule": r.get("schedule"),
                                    "shift_code": (
                                        provided_shift_code_by_key.get((ec, wd))
                                        if ec
                                        else None
                                    )
                                    or (
                                        provided_shift_code_by_key.get((ac, wd))
                                        if ac
                                        else None
                                    )
                                    or r.get("shift_code"),
                                }
                            )
                        except Exception:
                            continue

                    if pending_updates:
                        repo2 = ShiftAttendanceMainContent2Repository()

                        # Temporarily unlock imported rows so computed-field updates follow
                        # the same rules as normal rows, then lock again after done.
                        unlocked_ok = False
                        try:
                            n_unlock = repo2.update_import_locked_by_id(
                                pending_updates,
                                import_locked=0,
                            )
                            unlocked_ok = int(n_unlock or 0) > 0
                        except Exception:
                            # Best-effort: continue; we still can update with allow_import_locked.
                            repo2 = ShiftAttendanceMainContent2Repository()

                        n_updated = 0
                        try:
                            n_updated = repo2.update_computed_fields_by_id(
                                pending_updates,
                                allow_import_locked=False,
                            )

                            # If unlock failed silently or WHERE(import_locked=0) matches nothing,
                            # update 0 rows without raising. In that case, force-update.
                            if (not bool(unlocked_ok)) or int(n_updated or 0) <= 0:
                                repo2.update_computed_fields_by_id(
                                    pending_updates,
                                    allow_import_locked=True,
                                )
                        except Exception:
                            # Fallback: allow updating even if rows stayed locked.
                            repo2.update_computed_fields_by_id(
                                pending_updates,
                                allow_import_locked=True,
                            )

                        try:
                            repo2.update_import_locked_by_id(
                                pending_updates,
                                import_locked=1,
                            )
                        except Exception:
                            pass
        except Exception:
            # Do not fail the whole import if recompute fails; import itself succeeded.
            logger.exception("Không thể tính lại/cập nhật các cột tính toán sau import")

        extra = (
            f" | Bổ sung ngày thiếu: {inserted_missing}"
            if int(inserted_missing) > 0
            else ""
        )
        msg = f"Hoàn tất import: Thêm {inserted}, Cập nhập {updated}, Bỏ qua {skipped}, Lỗi {failed}.{extra}"
        return ImportShiftAttendanceResult(
            True,
            msg,
            inserted=inserted,
            updated=updated,
            skipped=skipped,
            failed=failed,
        )
